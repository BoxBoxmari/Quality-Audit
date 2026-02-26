#!/usr/bin/env python
"""Analyze XLSX output files for verification report."""

import sys
from collections import Counter

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

all_rows = []

for xlsx_path, doc_name in [
    (
        r"c:\Users\Admin\Downloads\Quality Audit (1)\Quality Audit\results\CJCGV-FS2018-EN- v2 _output.xlsx",
        "CJCGV",
    ),
    (
        r"c:\Users\Admin\Downloads\Quality Audit (1)\Quality Audit\results\CP Vietnam-FS2018-Consol-EN_output.xlsx",
        "CP",
    ),
]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb["Tổng hợp kiểm tra"]
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        row_dict["_doc"] = doc_name
        all_rows.append(row_dict)

# Status distribution
status_counts = Counter(r.get("Status Enum") for r in all_rows)
print("=== STATUS DISTRIBUTION ===")
for status, count in sorted(status_counts.items(), key=lambda x: -x[1]):
    print(f"{status}: {count}")
print(f"Total: {sum(status_counts.values())}")

# Total row method pivot for FAILs and WARNs
print()
print("=== TOTAL_ROW_METHOD PIVOT (FAIL/WARN/FAIL_TOOL_EXTRACT) ===")
fail_warn = [
    r for r in all_rows if r.get("Status Enum") in ("FAIL", "WARN", "FAIL_TOOL_EXTRACT")
]
method_status = Counter(
    (r.get("Total Row Method") or "EMPTY", r.get("Status Enum")) for r in fail_warn
)
for (method, status), count in sorted(method_status.items()):
    print(f"{method} + {status}: {count}")

# FAIL_TOOL_EXTRACT details
print()
print("=== FAIL_TOOL_EXTRACT DETAILS ===")
fail_extract = [r for r in all_rows if r.get("Status Enum") == "FAIL_TOOL_EXTRACT"]
for r in fail_extract:
    table_name = str(r.get("Tên bảng", ""))[:50]
    engine = r.get("Extractor Engine")
    qs = r.get("Quality Score")
    attempts = r.get("Engine Attempts")
    reason = r.get("Failure Reason Code")
    print(
        f"{r['_doc']} | {table_name} | Engine={engine} | QS={qs} | Attempts={attempts} | Reason={reason}"
    )

# Engine Attempts analysis
print()
print("=== ENGINE ATTEMPTS DISTRIBUTION ===")
attempts_counter = Counter(r.get("Engine Attempts") for r in all_rows)
for attempts, count in sorted(attempts_counter.items()):
    print(f"{attempts}: {count}")

# Check for render_first or libreoffice in attempts
print()
print("=== RENDER_FIRST / LIBREOFFICE CHECK ===")
for r in all_rows:
    attempts = str(r.get("Engine Attempts", "") or "")
    if "render" in attempts.lower() or "libre" in attempts.lower():
        print(f"{r['_doc']} | {r.get('Tên bảng', '')[:40]} | Attempts={attempts}")

# INFO/WARN details
print()
print("=== INFO/WARN DETAILS ===")
for r in all_rows:
    if r.get("Status Enum") in ("INFO", "WARN"):
        table_name = str(r.get("Tên bảng", ""))[:40]
        rule = r.get("Rule ID")
        method = r.get("Total Row Method")
        reason = r.get("Failure Reason Code")
        print(
            f"{r['Status Enum']} | {r['_doc']} | {table_name} | Rule={rule} | TRM={method} | Reason={reason}"
        )
