"""
Phase 6: Integration golden test — full pipeline on sample DOCX; assert summary sheet and Run metadata Per-Table Extraction columns.
"""

import pytest
from openpyxl import load_workbook

from quality_audit.core.cache_manager import AuditContext, LRUCacheManager
from quality_audit.io.excel_writer import ExcelWriter
from quality_audit.io.file_handler import FileHandler
from quality_audit.io.word_reader import AsyncWordReader
from quality_audit.services.audit_service import AuditService


class TestIntegrationGolden:
    """Run pipeline on sample_word_file; verify summary sheet and Run metadata columns."""

    @pytest.fixture
    def audit_service(self):
        cache_manager = LRUCacheManager(max_size=1000)
        context = AuditContext(cache=cache_manager)
        async_word_reader = AsyncWordReader(max_workers=4)
        return AuditService(
            context=context,
            async_word_reader=async_word_reader,
            excel_writer=ExcelWriter(),
            file_handler=FileHandler(),
        )

    @pytest.mark.asyncio
    async def test_summary_sheet_has_heading_classifier_assertions(
        self, audit_service, sample_word_file, tmp_path
    ):
        excel_path = tmp_path / "golden_output.xlsx"
        result = await audit_service.process_document_async(
            str(sample_word_file), str(excel_path)
        )
        assert result["success"] is True, result.get("error")
        assert excel_path.exists()

        wb = load_workbook(excel_path)
        assert "Tổng hợp kiểm tra" in wb.sheetnames
        ws_summary = wb["Tổng hợp kiểm tra"]
        assert ws_summary.max_row >= 2, (
            "Summary sheet should have header + at least one data row"
        )
        # Summary sheet: A=Tên bảng, B=Trạng thái kiểm tra, C=Status Enum (excel_writer write_summary_sheet)
        assert ws_summary.cell(row=1, column=3).value == "Status Enum"

        # Heading Source, Classifier Reason, Assertions Count are on Run metadata, Per-Table Extraction block (excel_writer)
        assert "Run metadata" in wb.sheetnames
        ws_run = wb["Run metadata"]
        header_row = None
        for r in range(1, min(ws_run.max_row + 1, 100)):
            if ws_run.cell(row=r, column=1).value == "Per-Table Extraction":
                header_row = r + 1
                break
        assert header_row is not None, (
            "Per-Table Extraction section not found on Run metadata"
        )
        assert ws_run.cell(row=header_row, column=3).value == "Heading Source"
        assert ws_run.cell(row=header_row, column=12).value == "Classifier Reason"
        assert ws_run.cell(row=header_row, column=14).value == "Assertions Count"
