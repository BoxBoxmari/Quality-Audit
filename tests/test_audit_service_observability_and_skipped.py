import pandas as pd
from openpyxl import load_workbook

from quality_audit.core.cache_manager import AuditContext, LRUCacheManager
from quality_audit.core.validators.base_validator import ValidationResult
from quality_audit.services.audit_service import AuditService


class TestAuditServiceObservabilityAndSkippedTables:
    def _make_service(self) -> AuditService:
        ctx = AuditContext(cache=LRUCacheManager(max_size=100))
        return AuditService(context=ctx)

    def test_skipped_footer_signature_table_produces_info_result(self, tmp_path):
        service = self._make_service()

        df = pd.DataFrame({"A": ["Signed by", "Director"], "B": ["", ""]})
        table_ctx = {"heading_source": "skipped"}

        result = service._validate_single_table(
            df, heading="SKIPPED_FOOTER_SIGNATURE", table_context=table_ctx
        )

        assert result.status_enum == "INFO"
        assert result.rule_id == "SKIPPED_FOOTER_SIGNATURE"
        assert "footer/signature" in result.context.get("reason", "")

    def test_validate_tables_enriches_context_and_logs(self, monkeypatch):
        def _get_service_flags():
            from quality_audit.config.feature_flags import (
                get_feature_flags as _global_get_flags,
            )

            flags = _global_get_flags().copy()
            flags["enable_big4_engine"] = False
            return flags

        monkeypatch.setattr(
            "quality_audit.services.audit_service.get_feature_flags",
            _get_service_flags,
        )
        service = self._make_service()

        df = pd.DataFrame(
            {
                "Code": ["10", "20", "Total"],
                "CY": [100.0, 200.0, 300.0],
                "PY": [90.0, 180.0, 270.0],
            }
        )
        table_ctx = {"heading_source": "table_first_row"}

        all_results = service._validate_tables([(df, "test heading", table_ctx)])

        assert len(all_results) == 1
        ctx = all_results[0]["context"]

        assert ctx.get("heading") == "test heading"
        assert ctx.get("table_index") == 0
        assert ctx.get("heading_source") == "table_first_row"
        # classifier_reason and scan_rows should be present in context even if values are None/0
        assert "classifier_reason" in ctx
        assert "scan_rows" in ctx
        assert "validator_type" in ctx
        assert "excluded_columns" in ctx
        assert "table_shape" in ctx

    def test_footer_artifacts_excluded_from_output(self, tmp_path):
        """Footer/signature tables are excluded from report output and telemetry counts them."""
        service = self._make_service()
        df_ok = pd.DataFrame(
            {"Code": ["10", "20"], "CY": [100.0, 200.0], "PY": [90.0, 180.0]}
        )
        df_footer = pd.DataFrame({"A": ["Signed by", "Director"], "B": ["", ""]})
        table_heading_pairs = [
            (df_ok, "Income Statement"),
            (df_footer, "SKIPPED_FOOTER_SIGNATURE"),
        ]
        results = service._validate_tables(table_heading_pairs)
        assert len(results) == 2
        results[1]["rule_id"] = "SKIPPED_FOOTER_SIGNATURE"
        results[1]["status_enum"] = "INFO"
        results[1]["context"] = results[1].get("context") or {}
        results[1]["context"]["table_id"] = "tbl_002_footer"

        excel_path = tmp_path / "out.xlsx"
        service._generate_report(
            table_heading_pairs, results, str(excel_path), telemetry=service.telemetry
        )

        wb = load_workbook(str(excel_path), read_only=False, data_only=True)
        telemetry_ws = wb["Run metadata"]
        assert telemetry_ws["A14"].value == "Skipped Footer/Signature Count"
        assert telemetry_ws["B14"].value == 1
        wb.close()

    def test_balance_sheet_not_skipped_by_numeric_gating_in_parity_mode(
        self, tmp_path, monkeypatch
    ):
        """
        In legacy parity mode, BalanceSheet tables must not be skipped by numeric-evidence gating.
        """

        def _get_factory_flags():
            return {
                "legacy_parity_mode": True,
                "routing_balance_sheet_gating_enabled": True,
                "routing_balance_sheet_gating_policy": "skip_no_numeric",
                "routing_balance_sheet_numeric_threshold": 0.25,
                "skip_footer_signature_tables": True,
            }

        def _get_service_flags():
            return {
                "metrics_exclude_footer_signature_artifacts": True,
            }

        monkeypatch.setattr(
            "quality_audit.core.validators.factory.get_feature_flags",
            _get_factory_flags,
        )
        monkeypatch.setattr(
            "quality_audit.services.audit_service.get_feature_flags",
            _get_service_flags,
        )

        service = self._make_service()

        # Create a BalanceSheet table with ASSETS/LIABILITIES keywords but text-only columns
        # (low numeric evidence score)
        df_bs_no_numeric = pd.DataFrame(
            {
                "Description": ["Total Assets", "Total Liabilities", "Equity"],
                "Notes": ["Narrative text only", "No figures", "Text column"],
            }
        )
        df_ok = pd.DataFrame(
            {"Code": ["10", "20"], "CY": [100.0, 200.0], "PY": [90.0, 180.0]}
        )

        table_heading_pairs = [
            (df_ok, "Income Statement"),
            (
                df_bs_no_numeric,
                "Balance Sheet",
            ),  # Use exact heading match to trigger BalanceSheet classification
        ]

        results = service._validate_tables(table_heading_pairs)
        assert len(results) == 2

        # In parity mode, this table must not be skipped via routing gate.
        bs_result = results[1]
        assert bs_result["rule_id"] != "SKIPPED_NO_NUMERIC_EVIDENCE"

        # Generate report and verify table is still included in output.
        excel_path = tmp_path / "out.xlsx"
        service._generate_report(table_heading_pairs, results, str(excel_path))

        wb = load_workbook(str(excel_path), read_only=False, data_only=True)
        sheet_names = wb.sheetnames
        assert len(sheet_names) >= 2
        wb.close()

    def test_ticket5_escape_hatch_disabled_in_parity_mode(self, monkeypatch):
        """Ticket-5 escape hatch must not force GenericTableValidator in parity mode."""

        monkeypatch.setattr(
            "quality_audit.services.audit_service.get_feature_flags",
            lambda: {"legacy_parity_mode": True},
        )

        service = self._make_service()
        service.context.set_last_classification_context(
            {
                "classifier_primary_type": "FS_BALANCE_SHEET",
                "classifier_confidence": 0.95,
            }
        )

        monkeypatch.setattr(
            "quality_audit.services.audit_service.ValidatorFactory.get_validator",
            lambda *_args, **_kwargs: (None, "SKIPPED_NO_NUMERIC_EVIDENCE"),
        )

        df = pd.DataFrame(
            {"Description": ["Assets", "Liabilities"], "Notes": ["n/a", "n/a"]}
        )
        result = service._validate_single_table(
            df,
            heading="Balance Sheet",
            table_context={"heading_source": "paragraph"},
        )

        assert result.rule_id == "NO_NUMERIC_EVIDENCE"

    def test_validate_single_table_injects_heading_attr_and_cleans_up(
        self, monkeypatch
    ):
        """Inject heading into df.attrs for validator and restore attrs after validation."""

        class _DummyValidator:
            def validate(self, table, _heading, table_context=None):
                captured = table.attrs.get("heading")
                return ValidationResult(
                    status="PASS",
                    status_enum="PASS",
                    context={"captured_heading_attr": captured},
                )

        monkeypatch.setattr(
            "quality_audit.services.audit_service.ValidatorFactory.get_validator",
            lambda *_args, **_kwargs: (_DummyValidator(), None),
        )
        monkeypatch.setattr(
            "quality_audit.services.audit_service.get_feature_flags",
            lambda: {"baseline_authoritative_default": False},
        )

        service = self._make_service()
        df = pd.DataFrame({"Code": ["10"], "CY": [100.0], "PY": [90.0]})
        assert "heading" not in df.attrs

        result = service._validate_single_table(
            df,
            heading="Injected Heading",
            table_context={"heading_source": "test"},
        )

        assert result.context.get("captured_heading_attr") == "Injected Heading"
        assert "heading" not in df.attrs
