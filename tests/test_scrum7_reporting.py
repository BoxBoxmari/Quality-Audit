import pandas as pd
import pytest
from openpyxl import Workbook

from quality_audit.core.validators.base_validator import BaseValidator, ValidationResult
from quality_audit.io.excel_writer import ExcelWriter


# Mock Validator for scoring testing
class MockValidator(BaseValidator):
    def validate(self, df, heading=None):
        return ValidationResult("TEST")


@pytest.fixture
def mock_results():
    return [
        ValidationResult(
            status="FAIL: Test failure",
            status_enum="FAIL",
            rule_id="MATH_EQ",
            severity="HIGH",
            confidence="HIGH",
            table_id="tbl_001",
            context={"heading": "Table 1", "validator_type": "GenericTableValidator"},
            evidence=[{"diff": 1000000, "label": "Row 1"}],
        ).to_dict(),
        ValidationResult(
            status="WARN: Test warning",
            status_enum="WARN",
            rule_id="MISSING_CODE_COL",
            severity="MEDIUM",
            confidence="MEDIUM",
            table_id="tbl_002",
            context={"heading": "Table 2", "validator_type": "GenericTableValidator"},
        ).to_dict(),
        ValidationResult(
            status="PASS: Test pass",
            status_enum="PASS",
            rule_id="MATH_EQ",
            table_id="tbl_003",
            context={"heading": "Table 3", "validator_type": "GenericTableValidator"},
        ).to_dict(),
    ]


def test_scoring_logic():
    validator = MockValidator()

    # Test Severity
    assert validator._calculate_severity("MATH_EQ", 2000000) == "HIGH"  # > 1M
    assert validator._calculate_severity("MATH_EQ", 500) == "LOW"  # < 1K
    assert (
        validator._calculate_severity("SKIPPED_MOVEMENT_TABLE", 0, is_skipped=True)
        == "LOW"
    )

    # Test Confidence
    assert validator._calculate_confidence("MATH_EQ") == "HIGH"
    assert validator._calculate_confidence("MATH_EQ", is_skipped=True) == "MEDIUM"


def test_executive_summary_creation(mock_results):
    writer = ExcelWriter()
    wb = Workbook()

    writer.write_executive_summary(wb, mock_results)

    ws = wb["Executive Summary"]
    assert ws["A1"].value == "EXECUTIVE SUMMARY"

    # Check metrics
    # Row 5: label/value pairs
    assert ws["A5"].value == "Total Tables"
    assert ws["B5"].value == 3
    assert ws["C5"].value == "Passed"
    assert ws["D5"].value == 1
    assert ws["E5"].value == "Failed"
    assert ws["F5"].value == 1
    assert ws["G5"].value == "Warnings"
    assert ws["H5"].value == 1

    # Check Overall Status
    assert "Overall Assessment" in ws["A7"].value
    assert ws["B7"].value == "RED"  # Has failures

    # Check Top 10 Table
    # Row 11 should be the failure
    assert ws.cell(row=11, column=1).value == "Table 1"  # Heading
    assert ws.cell(row=11, column=3).value == "HIGH"  # Severity
    assert ws.cell(row=11, column=4).value == 1000000  # Diff


def test_focus_list_creation(mock_results):
    writer = ExcelWriter()
    wb = Workbook()

    writer.write_focus_list(wb, mock_results)
    ws = wb["Focus List"]

    # Row 1 Headers - updated structure
    assert ws["A1"].value == "Table Name"
    assert ws["C1"].value == "Severity"
    assert ws["E1"].value == "Issue Description"

    # Row 2 (First finding - sorted by severity HIGH)
    assert ws.cell(row=2, column=1).value == "Table 1"  # Table Name
    assert ws.cell(row=2, column=3).value == "HIGH"  # Severity

    # Row 3 (Second finding - MEDIUM)
    assert ws.cell(row=3, column=1).value == "Table 2"  # Table Name
    assert ws.cell(row=3, column=3).value == "MEDIUM"  # Severity


def test_write_tables_consolidated_anchors():
    writer = ExcelWriter()
    wb = Workbook()

    df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
    pairs = [(df, "Table 1")]
    results = [ValidationResult("TEST", table_id="tbl_001").to_dict()]

    writer.write_tables_consolidated(wb, pairs, results)

    # Check Named Range
    assert "tbl_001" in wb.defined_names
    dn = wb.defined_names["tbl_001"]
    # Check it points to 'FS casting'!$A$1
    assert "FS casting" in dn.attr_text
