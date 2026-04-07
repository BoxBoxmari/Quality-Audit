#!/usr/bin/env python3
"""
Semantic parity harness for legacy-canonical workbook comparison.

Compares baseline/current XLSX by:
- sheet presence/order/name
- cell values
- comments
- fill color semantics (alpha-normalized RGB: FFxxxxxx == 00xxxxxx)
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Dict, Tuple

from openpyxl import load_workbook


def _cell_text(value) -> str:
    return "" if value is None else str(value)


def _normalize_rgb(rgb: str | None) -> str | None:
    if not rgb:
        return None
    rgb = rgb.upper()
    # Normalize ARGB to RGB semantics: treat FFxxxxxx and 00xxxxxx as same color.
    if len(rgb) == 8:
        return rgb[-6:]
    return rgb


def _fill_semantic(cell) -> Tuple[str | None, str | None]:
    fill_type = cell.fill.fill_type
    fg = cell.fill.fgColor
    rgb = None
    if fg is not None:
        rgb = _normalize_rgb(fg.rgb)
    return fill_type, rgb


def compare_semantic(baseline_path: Path, current_path: Path) -> Dict:
    wb_base = load_workbook(baseline_path)
    wb_cur = load_workbook(current_path)

    report: Dict = {
        "baseline": str(baseline_path),
        "current": str(current_path),
        "sheet_order_baseline": wb_base.sheetnames,
        "sheet_order_current": wb_cur.sheetnames,
        "missing_sheets": [s for s in wb_base.sheetnames if s not in wb_cur.sheetnames],
        "extra_sheets": [s for s in wb_cur.sheetnames if s not in wb_base.sheetnames],
        "sheet_metrics": {},
    }

    for sheet in sorted(set(wb_base.sheetnames) & set(wb_cur.sheetnames)):
        ws_b = wb_base[sheet]
        ws_c = wb_cur[sheet]
        max_row = max(ws_b.max_row, ws_c.max_row)
        max_col = max(ws_b.max_column, ws_c.max_column)

        value_mismatch = 0
        comment_mismatch = 0
        fill_mismatch = 0

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                cb = ws_b.cell(r, c)
                cc = ws_c.cell(r, c)

                if _cell_text(cb.value) != _cell_text(cc.value):
                    value_mismatch += 1

                if _cell_text(cb.comment.text if cb.comment else None) != _cell_text(
                    cc.comment.text if cc.comment else None
                ):
                    comment_mismatch += 1

                if _fill_semantic(cb) != _fill_semantic(cc):
                    fill_mismatch += 1

        report["sheet_metrics"][sheet] = {
            "value_mismatches": value_mismatch,
            "comment_mismatches": comment_mismatch,
            "fill_semantic_mismatches": fill_mismatch,
            "dimensions_baseline": [ws_b.max_row, ws_b.max_column],
            "dimensions_current": [ws_c.max_row, ws_c.max_column],
        }

    report["pass"] = (
        not report["missing_sheets"]
        and not report["extra_sheets"]
        and all(
            m["value_mismatches"] == 0
            and m["comment_mismatches"] == 0
            and m["fill_semantic_mismatches"] == 0
            for m in report["sheet_metrics"].values()
        )
    )
    return report


def main() -> int:
    parser = argparse.ArgumentParser(description="Run semantic XLSX parity harness.")
    parser.add_argument(
        "--baseline",
        type=Path,
        default=Path(r"C:\Users\Admin\Downloads\ABC Company-30Jun26VAS-EN.xlsx"),
    )
    parser.add_argument(
        "--current",
        type=Path,
        default=Path(
            r"C:\Users\Admin\Downloads\Quality Audit Tool\results\ABC Company-30Jun26VAS-EN.current.xlsx"
        ),
    )
    parser.add_argument("--json-out", type=Path, default=None)
    args = parser.parse_args()

    if not args.baseline.exists():
        raise FileNotFoundError(f"Baseline XLSX not found: {args.baseline}")
    if not args.current.exists():
        raise FileNotFoundError(f"Current XLSX not found: {args.current}")

    report = compare_semantic(args.baseline, args.current)
    payload = json.dumps(report, ensure_ascii=False, indent=2)
    print(payload)

    if args.json_out:
        args.json_out.parent.mkdir(parents=True, exist_ok=True)
        args.json_out.write_text(payload, encoding="utf-8")

    return 0 if report["pass"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
