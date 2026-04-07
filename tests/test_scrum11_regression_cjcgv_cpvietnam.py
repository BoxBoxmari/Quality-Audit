"""
SCRUM-11 Regression Tests: Golden file tests for CJCGV and CP Vietnam datasets.

These tests ensure that after rollback to pre-SCRUM-9 baseline, the tool:
1. Runs end-to-end without crashing
2. Produces consistent output structure
3. Maintains SCRUM-5→8 features (telemetry, Executive Summary, Focus List, hyperlinks)
"""

from pathlib import Path

import pytest

from quality_audit.services.audit_service import AuditService


class TestSCRUM11RegressionCJCGVCPVietnam:
    """
    Regression tests for SCRUM-11: Restore baseline pre-SCRUM-9.

    Tests ensure CJCGV and CP Vietnam datasets process successfully
    and produce expected output structure.
    """

    @pytest.fixture
    def audit_service(self):
        """Create AuditService instance for testing."""
        return AuditService()

    @pytest.fixture
    def test_data_dir(self):
        """Get test data directory path."""
        base_dir = Path(__file__).parent.parent
        test_data = base_dir / "test_data"
        if not test_data.exists():
            test_data = base_dir / "data"
        return test_data

    def test_cjcgv_fs2018_runs_without_crash(
        self, audit_service, test_data_dir, tmp_path
    ):
        """
        Test: CJCGV FS2018 document processes end-to-end without crashing.

        Expected: No exceptions, output Excel file created with expected sheets.
        """
        word_file = test_data_dir / "CJCGV-FS2018-EN- v2.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cjcgv_fs2018_output.xlsx"

        # Should not raise any exception
        result = audit_service.audit_document(str(word_file), str(excel_output))

        # Verify success
        assert result["success"] is True, f"Audit failed: {result.get('error')}"
        assert result["tables_processed"] > 0, "No tables were processed"
        assert excel_output.exists(), "Output Excel file was not created"

        # Verify output structure (basic checks)
        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        # SCRUM-7: Executive Summary and Focus List should exist
        assert "Executive Summary" in wb.sheetnames, "Executive Summary sheet missing"
        assert "Focus List" in wb.sheetnames, "Focus List sheet missing"

        # SCRUM-6: Run metadata sheet should exist (telemetry)
        assert "Run metadata" in wb.sheetnames, "Run metadata sheet missing (telemetry)"

        # FS casting sheet should exist
        assert "FS casting" in wb.sheetnames, "FS casting sheet missing"

        wb.close()

    def test_cp_vietnam_fs2018_consol_runs_without_crash(
        self, audit_service, test_data_dir, tmp_path
    ):
        """
        Test: CP Vietnam FS2018 Consol document processes end-to-end without crashing.

        Expected: No exceptions, output Excel file created with expected sheets.
        """
        word_file = test_data_dir / "CP Vietnam-FS2018-Consol-EN.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cp_vietnam_fs2018_output.xlsx"

        # Should not raise any exception
        result = audit_service.audit_document(str(word_file), str(excel_output))

        # Verify success
        assert result["success"] is True, f"Audit failed: {result.get('error')}"
        assert result["tables_processed"] > 0, "No tables were processed"
        assert excel_output.exists(), "Output Excel file was not created"

        # Verify output structure (basic checks)
        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        # SCRUM-7: Executive Summary and Focus List should exist
        assert "Executive Summary" in wb.sheetnames, "Executive Summary sheet missing"
        assert "Focus List" in wb.sheetnames, "Focus List sheet missing"

        # SCRUM-6: Run metadata sheet should exist (telemetry)
        assert "Run metadata" in wb.sheetnames, "Run metadata sheet missing (telemetry)"

        # FS casting sheet should exist
        assert "FS casting" in wb.sheetnames, "FS casting sheet missing"

        wb.close()

    def test_cjcgv_telemetry_populated(self, audit_service, test_data_dir, tmp_path):
        """
        Test: Telemetry data is populated in Run metadata sheet (SCRUM-6).

        Expected: Run metadata sheet contains tool_version, git_commit_hash, run_timestamp.
        """
        word_file = test_data_dir / "CJCGV-FS2018-EN- v2.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cjcgv_telemetry_test.xlsx"

        result = audit_service.audit_document(str(word_file), str(excel_output))
        assert result["success"] is True

        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        if "Run metadata" not in wb.sheetnames:
            wb.close()
            pytest.skip("Run metadata sheet not created")

        ws = wb["Run metadata"]

        # Check build identification fields (SCRUM-6)
        tool_version = ws["B4"].value
        git_commit = ws["B5"].value
        run_timestamp = ws["B6"].value

        assert tool_version is not None, "Tool version should be populated"
        assert git_commit is not None, "Git commit should be populated (or 'unknown')"
        assert run_timestamp is not None, "Run timestamp should be populated"

        # Check performance metrics
        total_runtime = ws["B9"].value
        table_count = ws["B10"].value

        assert total_runtime is not None, "Total runtime should be populated"
        assert table_count is not None and table_count > 0, "Table count should be > 0"

        wb.close()

    def test_focus_list_has_hyperlinks(self, audit_service, test_data_dir, tmp_path):
        """
        Test: Focus List contains hyperlinks to FS casting (SCRUM-8).

        Expected: Jump column (G) contains hyperlinks or "(missing anchor)" fallback.
        """
        word_file = test_data_dir / "CJCGV-FS2018-EN- v2.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cjcgv_hyperlinks_test.xlsx"

        result = audit_service.audit_document(str(word_file), str(excel_output))
        assert result["success"] is True

        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        if "Focus List" not in wb.sheetnames:
            wb.close()
            pytest.skip("Focus List sheet not created")

        ws = wb["Focus List"]

        # Check if there are any findings (FAIL/WARN rows)
        has_findings = False
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=2).value  # Status column
            if status in ["FAIL", "WARN"]:
                has_findings = True
                # Check Jump column (G, column 7)
                jump_cell = ws.cell(row=row_idx, column=7)

                # Should have hyperlink or "(missing anchor)" text
                has_link = jump_cell.hyperlink is not None
                has_fallback = jump_cell.value == "(missing anchor)"

                assert (
                    has_link or has_fallback
                ), f"Row {row_idx} should have hyperlink or fallback text"

                if has_link:
                    # Verify hyperlink format
                    target = jump_cell.hyperlink.target
                    assert target.startswith(
                        "#"
                    ), f"Hyperlink should be internal: {target}"
                    assert (
                        "'FS casting'" in target or "FS casting" in target
                    ), f"Hyperlink should point to FS casting: {target}"

        if not has_findings:
            pytest.skip("No FAIL/WARN findings to test hyperlinks")

        wb.close()

    def test_severity_not_all_medium(self, audit_service, test_data_dir, tmp_path):
        """
        Test: Severity values are not all defaulting to MEDIUM (SCRUM-8).

        Expected: At least some HIGH or LOW severity values if there are findings.
        """
        word_file = test_data_dir / "CJCGV-FS2018-EN- v2.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cjcgv_severity_test.xlsx"

        result = audit_service.audit_document(str(word_file), str(excel_output))
        assert result["success"] is True

        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        if "Focus List" not in wb.sheetnames:
            wb.close()
            pytest.skip("Focus List sheet not created")

        ws = wb["Focus List"]

        # Collect severity values
        severities = []
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=2).value
            if status in ["FAIL", "WARN"]:
                severity = ws.cell(row=row_idx, column=3).value  # Severity column
                if severity:
                    severities.append(severity)

        if not severities:
            pytest.skip("No findings to test severity distribution")

        # Should have at least one non-MEDIUM severity
        non_medium = [s for s in severities if s != "MEDIUM"]
        assert (
            len(non_medium) > 0
        ), f"All severities are MEDIUM (defaulting issue). Found: {set(severities)}"

    def test_root_cause_not_all_general(self, audit_service, test_data_dir, tmp_path):
        """
        Test: Root cause values are not all defaulting to "general" (SCRUM-8).

        Expected: At least some specific root causes (calculation, mapping, subtotal, etc.).
        """
        word_file = test_data_dir / "CJCGV-FS2018-EN- v2.docx"
        if not word_file.exists():
            pytest.skip(f"Test data not found: {word_file}")

        excel_output = tmp_path / "cjcgv_root_cause_test.xlsx"

        result = audit_service.audit_document(str(word_file), str(excel_output))
        assert result["success"] is True

        import openpyxl

        wb = openpyxl.load_workbook(excel_output)

        if "Focus List" not in wb.sheetnames:
            wb.close()
            pytest.skip("Focus List sheet not created")

        ws = wb["Focus List"]

        # Collect root cause values
        root_causes = []
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=2).value
            if status in ["FAIL", "WARN"]:
                root_cause = ws.cell(row=row_idx, column=10).value  # Root Cause column
                if root_cause:
                    root_causes.append(root_cause)

        if not root_causes:
            pytest.skip("No findings to test root cause distribution")

        # Should have at least one non-"general" root cause
        non_general = [rc for rc in root_causes if rc and rc.lower() != "general"]
        assert (
            len(non_general) > 0
        ), f"All root causes are 'general' (defaulting issue). Found: {set(root_causes)}"
