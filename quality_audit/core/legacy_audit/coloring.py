"""
Legacy Excel coloring/cell-mark semantics.
"""

from openpyxl.styles import Alignment, Font, PatternFill

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
BLUE_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
INFO_FILL = PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid")

GREEN_FONT = Font(color="32CD32")
RED_FONT = Font(color="FF0000")
RIGHT_ALIGN = Alignment(horizontal="right")
