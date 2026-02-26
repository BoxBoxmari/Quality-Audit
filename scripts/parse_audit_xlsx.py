"""Parse audit XLSX files to extract status distribution and non-PASS tables."""

import json
import sys
from pathlib import Path

# Force UTF-8 output
sys.stdout.reconfigure(encoding="utf-8")

import openpyxl  # noqa: E402


def parse_xlsx(filepath: Path) -> dict:
    """Parse a single XLSX file and extract summary data."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = {
        "file": filepath.name,
        "sheets": wb.sheetnames,
        "summary_data": [],
        "focus_list_data": [],
    }

    # Target the exact Vietnamese summary sheet "Tổng hợp kiểm tra"
    summary_sheet = None
    for name in wb.sheetnames:
        if "Tổng hợp kiểm tra" in name or "Tong hop kiem tra" in name:
            summary_sheet = wb[name]
            result["summary_sheet_name"] = name
            break

    # Fallback if not found
    if not summary_sheet:
        for name in wb.sheetnames:
            if "summary" in name.lower() and "executive" not in name.lower():
                summary_sheet = wb[name]
                result["summary_sheet_name"] = name
                break

    if summary_sheet:
        # Read headers from first row
        headers = []
        for cell in summary_sheet[1]:
            headers.append(str(cell.value) if cell.value else "")
        result["summary_headers"] = headers

        # Read data rows
        for row_num, row in enumerate(
            summary_sheet.iter_rows(min_row=2, max_row=200), start=2
        ):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_data[headers[i]] = cell.value
            if any(v for v in row_data.values()):  # Skip empty rows
                row_data["_row_num"] = row_num
                result["summary_data"].append(row_data)

    # Try to find Focus List sheet
    for name in wb.sheetnames:
        if "focus" in name.lower():
            focus_sheet = wb[name]
            result["focus_sheet_name"] = name

            # Read headers
            headers = []
            for cell in focus_sheet[1]:
                headers.append(str(cell.value) if cell.value else "")
            result["focus_headers"] = headers

            # Read data rows
            for row_num, row in enumerate(
                focus_sheet.iter_rows(min_row=2, max_row=200), start=2
            ):
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers) and headers[i]:
                        row_data[headers[i]] = cell.value
                if any(v for v in row_data.values()):
                    row_data["_row_num"] = row_num
                    result["focus_list_data"].append(row_data)
            break

    wb.close()
    return result


def analyze_status_distribution(data: list) -> dict:
    """Analyze status distribution from summary data."""
    status_counts = {}
    non_pass_tables = []

    for row in data:
        # Find status column - could be "Status Enum", "Status", "Trạng thái"
        status = None
        for key in ["Status Enum", "Status", "Trạng thái", "status_enum"]:
            if key in row and row[key]:
                status = str(row[key]).strip().upper()
                break

        if status:
            status_counts[status] = status_counts.get(status, 0) + 1
            if status != "PASS":
                non_pass_tables.append(
                    {
                        "row": row.get("_row_num"),
                        "table_id": row.get("Table ID", row.get("table_id", "?")),
                        "status": status,
                        "validator": row.get(
                            "Validator Type", row.get("validator_type", "?")
                        ),
                        "failure_reason": row.get(
                            "Failure Reason", row.get("failure_reason", "")
                        ),
                        "quality_score": row.get(
                            "Quality Score", row.get("quality_score", "")
                        ),
                    }
                )

    return {
        "status_distribution": status_counts,
        "total": sum(status_counts.values()),
        "non_pass_tables": non_pass_tables,
    }


def main():
    results_dir = Path("results")
    xlsx_files = list(results_dir.glob("*_output.xlsx"))

    all_results = []
    combined_non_pass = []
    combined_status = {}

    for xlsx_file in xlsx_files:
        print(f"\n=== Processing: {xlsx_file.name} ===")
        data = parse_xlsx(xlsx_file)

        print(f"Sheets: {data['sheets']}")
        if "summary_sheet_name" in data:
            print(f"Summary sheet: {data['summary_sheet_name']}")
            print(f"Summary headers: {data.get('summary_headers', [])[:10]}...")
            print(f"Summary rows: {len(data['summary_data'])}")

            analysis = analyze_status_distribution(data["summary_data"])
            print(f"Status distribution: {analysis['status_distribution']}")
            print(f"Total tables: {analysis['total']}")
            print(f"Non-PASS count: {len(analysis['non_pass_tables'])}")

            # Combine results
            for status, count in analysis["status_distribution"].items():
                combined_status[status] = combined_status.get(status, 0) + count
            for item in analysis["non_pass_tables"]:
                item["source_file"] = xlsx_file.name
                combined_non_pass.append(item)

        if "focus_sheet_name" in data:
            print(f"Focus sheet: {data['focus_sheet_name']}")
            print(f"Focus rows: {len(data['focus_list_data'])}")

        all_results.append(data)

    print("\n" + "=" * 60)
    print("COMBINED RESULTS")
    print("=" * 60)
    print(f"Total status distribution: {combined_status}")
    print(f"Total tables: {sum(combined_status.values())}")
    print(f"\nNon-PASS tables ({len(combined_non_pass)}):")
    for item in combined_non_pass:
        print(
            f"  - {item['table_id']}: {item['status']} | {item['validator']} | {item.get('failure_reason', '')[:50]}"
        )

    # Save detailed results
    output = {
        "combined_status": combined_status,
        "total_tables": sum(combined_status.values()),
        "non_pass_tables": combined_non_pass,
    }
    with open("results/audit_analysis.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    print("\nSaved detailed results to results/audit_analysis.json")


if __name__ == "__main__":
    main()
