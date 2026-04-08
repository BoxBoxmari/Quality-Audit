"""
Integration tests for AuditService.
"""

import asyncio

import pytest
from openpyxl import load_workbook

from quality_audit.core.cache_manager import (
    AuditContext,
    LRUCacheManager,
    cross_check_cache,
    cross_check_marks,
)
from quality_audit.services.audit_service import AuditService


class TestAuditServiceIntegration:
    """Integration tests for complete audit workflow."""

    @pytest.fixture
    def audit_context(self):
        """Create audit context for testing."""
        cache = LRUCacheManager(max_size=100)
        return AuditContext(cache=cache)

    @pytest.fixture
    def audit_service(self, audit_context):
        """Create audit service instance."""
        return AuditService(context=audit_context)

    def test_full_audit_workflow(self, audit_service, sample_word_file, tmp_path):
        """
        Test complete audit workflow from Word to Excel.

        This test verifies the entire pipeline:
        1. Word document reading
        2. Table validation
        3. Excel report generation
        """
        excel_path = tmp_path / "output.xlsx"

        result = audit_service.audit_document(sample_word_file, str(excel_path))

        assert result["success"] is True
        assert result["tables_processed"] > 0
        assert excel_path.exists()

        # Verify Excel content
        wb = load_workbook(excel_path)
        assert "Tổng hợp kiểm tra" in wb.sheetnames

    def test_audit_service_with_invalid_file(self, audit_service, tmp_path):
        """Test audit service with invalid file path."""
        invalid_path = tmp_path / "nonexistent.docx"
        excel_path = tmp_path / "output.xlsx"

        result = audit_service.audit_document(str(invalid_path), str(excel_path))

        assert result["success"] is False
        assert "error" in result
        assert result["tables_processed"] == 0

    def test_audit_service_context_isolation(self, tmp_path):
        """Test that different audit contexts are isolated."""
        context1 = AuditContext(cache=LRUCacheManager(max_size=100))
        context2 = AuditContext(cache=LRUCacheManager(max_size=100))

        service1 = AuditService(context=context1)
        service2 = AuditService(context=context2)

        # Verify contexts are separate
        assert service1.context is not service2.context
        assert service1.context.cache is not service2.context.cache

    def test_repeated_sync_runs_reset_legacy_and_global_state(
        self, tmp_path, monkeypatch
    ):
        """Same service instance should be deterministic across repeated sync runs."""
        from quality_audit.services import audit_service as audit_service_module

        class FakeLegacyMain:
            def __init__(self):
                self.BSPL_cross_check_cache = {}
                self.BSPL_cross_check_mark = []

            def read_word_tables_with_headings(self, _word_path):
                return [("table", "Heading A")]

            def check_table_total(self, _table, _heading):
                old = self.BSPL_cross_check_cache.get("run_count", 0)
                self.BSPL_cross_check_cache["run_count"] = old + 1
                self.BSPL_cross_check_mark.append("marked")
                return {"status": f"run={self.BSPL_cross_check_cache['run_count']}"}

            def write_table_sheet(self, wb, _pairs, _results):
                ws = wb.active
                ws.title = "Tổng hợp kiểm tra"
                ws.append(["ok"])
                return {}

            def write_summary_sheet(self, *_args, **_kwargs):
                return None

        fake_legacy = FakeLegacyMain()
        monkeypatch.setattr(
            audit_service_module, "_load_legacy_main_module", lambda: fake_legacy
        )

        service = AuditService(
            context=AuditContext(cache=LRUCacheManager(max_size=100))
        )
        monkeypatch.setattr(service.file_handler, "validate_path", lambda _path: True)
        monkeypatch.setattr(
            service.file_handler, "validate_docx_safety", lambda _path: True
        )

        # Seed stale state to verify reset
        cross_check_cache.set("stale", (9.0, 8.0))
        cross_check_marks.add("stale")
        fake_legacy.BSPL_cross_check_cache["stale"] = (7.0, 6.0)
        fake_legacy.BSPL_cross_check_mark.append("stale")

        word_path = tmp_path / "dummy.docx"
        word_path.write_text("dummy")
        excel_1 = tmp_path / "out_1.xlsx"
        excel_2 = tmp_path / "out_2.xlsx"

        result_1 = service.audit_document(str(word_path), str(excel_1))
        result_2 = service.audit_document(str(word_path), str(excel_2))

        assert result_1["success"] is True
        assert result_2["success"] is True
        assert result_1["results"] == result_2["results"]
        assert result_1["results"][0]["status"] == "run=1"
        assert result_2["results"][0]["status"] == "run=1"
        assert "stale" not in cross_check_cache
        assert "stale" not in cross_check_marks
        assert "stale" not in fake_legacy.BSPL_cross_check_cache

    def test_repeated_async_runs_reset_state(self, tmp_path, monkeypatch):
        """Async shell path should remain deterministic across repeated runs."""
        from quality_audit.services import audit_service as audit_service_module

        class FakeLegacyMain:
            def __init__(self):
                self.BSPL_cross_check_cache = {}
                self.BSPL_cross_check_mark = []

            def read_word_tables_with_headings(self, _word_path):
                return [("table", "Heading A")]

            def check_table_total(self, _table, _heading):
                old = self.BSPL_cross_check_cache.get("run_count", 0)
                self.BSPL_cross_check_cache["run_count"] = old + 1
                return {"status": f"run={self.BSPL_cross_check_cache['run_count']}"}

            def write_table_sheet(self, wb, _pairs, _results):
                ws = wb.active
                ws.title = "Tổng hợp kiểm tra"
                ws.append(["ok"])
                return {}

            def write_summary_sheet(self, *_args, **_kwargs):
                return None

        fake_legacy = FakeLegacyMain()
        monkeypatch.setattr(
            audit_service_module, "_load_legacy_main_module", lambda: fake_legacy
        )

        service = AuditService(
            context=AuditContext(cache=LRUCacheManager(max_size=100))
        )
        monkeypatch.setattr(service.file_handler, "validate_path", lambda _path: True)
        monkeypatch.setattr(
            service.file_handler, "validate_docx_safety", lambda _path: True
        )

        word_path = tmp_path / "dummy_async.docx"
        word_path.write_text("dummy")
        excel_1 = tmp_path / "async_1.xlsx"
        excel_2 = tmp_path / "async_2.xlsx"

        result_1 = asyncio.run(
            service.process_document_async(str(word_path), str(excel_1))
        )
        result_2 = asyncio.run(
            service.process_document_async(str(word_path), str(excel_2))
        )

        assert result_1["success"] is True
        assert result_2["success"] is True
        assert result_1["results"] == result_2["results"]
        assert result_1["results"][0]["status"] == "run=1"
        assert result_2["results"][0]["status"] == "run=1"
