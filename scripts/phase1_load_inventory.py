"""
Phase 1.1: Load 2 output xlsx + log -> table inventory.

Reads results/CP Vietnam-FS2018-Consol-EN_output.xlsx and
results/CJCGV-FS2018-EN- v2 _output.xlsx; builds a table inventory with
table_id, heading, table_type, status_enum, status_category, rule_id, assertions_count.
"""

from __future__ import annotations

import contextlib
import re
import sys
from pathlib import Path

# Force UTF-8
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import openpyxl
except ImportError:
    print("openpyxl required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Project root
REPO_ROOT = Path(__file__).resolve().parents[1]
RESULTS_DIR = REPO_ROOT / "results"

# Default output xlsx paths
DEFAULT_XLSX = [
    RESULTS_DIR / "CP Vietnam-FS2018-Consol-EN_output.xlsx",
    RESULTS_DIR / "CJCGV-FS2018-EN- v2 _output.xlsx",
]

SUMMARY_SHEET = "Tổng hợp kiểm tra"
RUN_METADATA_SHEET = "Run metadata"
FS_CASTING_SHEET = "FS casting"

# Per-Table Extraction column indices (1-based in Excel)
COL_TABLE_INDEX = 1
COL_HEADING = 2
COL_CLASSIFIER_TYPE = 10
COL_ASSERTIONS_COUNT = 14


def _parse_table_id_from_fs_header(cell_value: str | None) -> str | None:
    """Parse 'raw_name [table_id]' -> table_id."""
    if not cell_value or " [" not in str(cell_value):
        return None
    s = str(cell_value).strip()
    match = re.search(r"\s\[([^\]]+)\]\s*$", s)
    if match:
        return match.group(1).strip()
    return None


def _read_summary_sheet(ws) -> list[dict]:
    """Read summary sheet; return list of row dicts (one per table)."""
    rows = []
    header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        return rows
    headers = [str(c.value or "").strip() for c in header_row]
    for row in ws.iter_rows(min_row=2, max_row=500):
        values = [c.value for c in row]
        if not any(v is not None and str(v).strip() for v in values):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(values) and h:
                row_dict[h] = values[i]
        row_dict["_excel_row"] = row[0].row
        rows.append(row_dict)
    return rows


def _read_run_metadata_per_table(ws) -> list[dict]:
    """Locate 'Per-Table Extraction' block and read table rows."""
    found_start = False
    header_row_idx = 0
    for row in ws.iter_rows(min_row=1, max_row=300):
        first_cell = row[0].value if row else None
        if first_cell and "Per-Table Extraction" in str(first_cell):
            found_start = True
            header_row_idx = row[0].row
            break
    if not found_start or not header_row_idx:
        return []

    # Next row is headers
    headers = []
    for cell in ws[header_row_idx + 1]:
        headers.append(str(cell.value or "").strip())
    # Data rows
    data = []
    for r in range(header_row_idx + 2, min(header_row_idx + 200, ws.max_row + 1)):
        row_vals = []
        for c in range(1, 20):
            cell = ws.cell(row=r, column=c)
            row_vals.append(cell.value)
        if not any(v is not None for v in row_vals):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if i < len(row_vals) and h:
                row_dict[h] = row_vals[i]
        row_dict["_row"] = r
        data.append(row_dict)
    return data


def _read_fs_casting_table_ids(ws) -> list[str | None]:
    """Scan FS casting column A for '... [table_id]' lines; return table_id list in order."""
    table_ids = []
    for row in ws.iter_rows(min_row=1, max_row=2000, max_col=1):
        cell = row[0]
        val = cell.value
        tid = _parse_table_id_from_fs_header(val)
        if tid is not None:
            table_ids.append(tid)
    return table_ids


def load_one_xlsx(filepath: Path) -> list[dict]:
    """Load one *_output.xlsx and return list of table inventory rows."""
    if not filepath.exists():
        return []

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    file_name = filepath.name

    # Summary sheet
    summary_ws = None
    for name in wb.sheetnames:
        if name == SUMMARY_SHEET or "Tổng hợp" in name:
            summary_ws = wb[name]
            break
    if not summary_ws:
        wb.close()
        return []

    summary_rows = _read_summary_sheet(summary_ws)

    # Run metadata -> Per-Table Extraction (Classifier Type, Assertions Count)
    run_ws = wb[RUN_METADATA_SHEET] if RUN_METADATA_SHEET in wb.sheetnames else None
    per_table = _read_run_metadata_per_table(run_ws) if run_ws else []

    # FS casting -> table_id per block
    fs_ws = wb[FS_CASTING_SHEET] if FS_CASTING_SHEET in wb.sheetnames else None
    table_ids = _read_fs_casting_table_ids(fs_ws) if fs_ws else []

    wb.close()

    # Align by index: summary_rows[i], per_table[i], table_ids[i]
    n = max(len(summary_rows), len(per_table), len(table_ids))
    inventory = []
    for i in range(n):
        sr = summary_rows[i] if i < len(summary_rows) else {}
        pt = per_table[i] if i < len(per_table) else {}
        table_id = table_ids[i] if i < len(table_ids) else None
        heading = sr.get("Tên bảng") or pt.get("Heading") or ""
        if isinstance(heading, float):
            heading = "" if heading != heading else str(int(heading))
        else:
            heading = str(heading or "").strip()
        status_enum = str(sr.get("Status Enum") or "").strip()
        status_category = str(sr.get("Status Category") or "").strip()
        rule_id = str(sr.get("Rule ID") or "").strip()
        validator_type = str(sr.get("Validator Type") or "").strip()
        failure_reason_code = str(sr.get("Failure Reason Code") or "").strip()
        table_type = str(pt.get("Classifier Type") or "").strip()
        # Normalize classifier value to enum name for T6 (fs_income_statement -> FS_INCOME_STATEMENT)
        _ft = (table_type or "").lower()
        if _ft == "fs_balance_sheet":
            table_type = "FS_BALANCE_SHEET"
        elif _ft == "fs_income_statement":
            table_type = "FS_INCOME_STATEMENT"
        elif _ft == "fs_cash_flow":
            table_type = "FS_CASH_FLOW"
        assertions_count = pt.get("Assertions Count")
        if assertions_count is not None and isinstance(assertions_count, (int, float)):
            with contextlib.suppress(ValueError, TypeError):
                assertions_count = int(assertions_count)
        else:
            assertions_count = None

        inventory.append(
            {
                "file_name": file_name,
                "table_index": i + 1,
                "table_id": table_id or "",
                "heading": heading,
                "table_type": table_type,
                "status_enum": status_enum,
                "status_category": status_category,
                "rule_id": rule_id,
                "validator_type": validator_type,
                "failure_reason_code": failure_reason_code,
                "assertions_count": assertions_count,
            }
        )
    return inventory


def load_inventory(xlsx_paths: list[Path] | None = None) -> list[dict]:
    """Load all given xlsx and return combined table inventory."""
    paths = xlsx_paths or DEFAULT_XLSX
    combined = []
    for path in paths:
        combined.extend(load_one_xlsx(Path(path)))
    return combined


def main() -> int:
    """CLI: load inventory and print as CSV to stdout."""
    inventory = load_inventory()
    if not inventory:
        print("No tables loaded. Check paths.", file=sys.stderr)
        return 1

    # CSV header
    keys = [
        "file_name",
        "table_index",
        "table_id",
        "heading",
        "table_type",
        "status_enum",
        "status_category",
        "rule_id",
        "validator_type",
        "failure_reason_code",
        "assertions_count",
    ]
    print(",".join(keys))
    for row in inventory:
        vals = []
        for k in keys:
            v = row.get(k, "")
            if v is None:
                v = ""
            v = str(v).replace('"', '""')
            if "," in v or "\n" in v or '"' in v:
                v = f'"{v}"'
            vals.append(v)
        print(",".join(vals))
    return 0


if __name__ == "__main__":
    sys.exit(main())
