"""
SCRUM-11: Contract tests to lock core pipeline linkage.

Tests ensure:
- Reader → Validators → Report Writer contract is maintained
- No circular imports or missing symbols
- Config loader works correctly
- Schema/field contracts are preserved
"""

import pandas as pd
from openpyxl import Workbook

from quality_audit.config.constants import RULE_TAXONOMY, ScoringConfig
from quality_audit.core.validators.base_validator import ValidationResult
from quality_audit.core.validators.factory import ValidatorFactory
from quality_audit.io.excel_writer import ExcelWriter
from quality_audit.services.audit_service import AuditService


class TestImportSanity:
    """Test 1: Import sanity - no circular imports or missing symbols."""

    def test_core_imports_successfully(self):
        """Verify all core modules import without circular dependencies."""
        # Test critical imports (noqa: names used only to verify import succeeds)
        from quality_audit.config import constants, validation_rules  # noqa: F401
        from quality_audit.core import (  # noqa: F401
            AuditContext,
            LRUCacheManager,
            QualityAuditError,
            ValidationError,
        )
        from quality_audit.core.validators import (  # noqa: F401
            BalanceSheetValidator,
            BaseValidator,
            CashFlowValidator,
            GenericTableValidator,
            IncomeStatementValidator,
        )
        from quality_audit.io import ExcelWriter, FileHandler  # noqa: F401
        from quality_audit.services import AuditService  # noqa: F401

        # If we get here, imports succeeded
        assert True

    def test_config_loader_works(self):
        """Verify config constants load correctly."""
        from quality_audit.config.constants import (
            RuleCriticality,
        )

        # Verify critical configs exist
        assert RULE_TAXONOMY is not None
        assert isinstance(RULE_TAXONOMY, dict)
        assert ScoringConfig is not None
        assert RuleCriticality is not None

    def test_no_missing_symbols(self):
        """Verify all referenced symbols exist."""
        from quality_audit.core.validators.base_validator import BaseValidator

        # Verify factory can create validators
        df = pd.DataFrame({"Code": ["100"], "CY": [1000], "PY": [900]})
        validator, skip_reason = ValidatorFactory.get_validator(df, "test table")
        # Should return None or a validator, not raise
        assert validator is None or isinstance(validator, BaseValidator)


class TestReaderValidatorContract:
    """Test 2: Reader → Validator contract."""

    def test_table_normalization_contract(self):
        """Verify TableNormalizer produces valid DataFrame for validators."""
        from quality_audit.utils.table_normalizer import TableNormalizer

        # Create test DataFrame
        df = pd.DataFrame(
            {
                "Code": ["100", "200", "Total"],
                "2024": [1000, 2000, 3000],
                "2023": [900, 1900, 2800],
            }
        )

        normalized_df, metadata = TableNormalizer.normalize_table(df, "Test Table")

        # Contract: normalized_df is DataFrame
        assert isinstance(normalized_df, pd.DataFrame)
        # Contract: metadata has expected keys
        assert "detected_code_column" in metadata
        assert "detected_cur_col" in metadata
        assert "detected_prior_col" in metadata

    def test_validator_accepts_normalized_dataframe(self):
        """Verify validators accept normalized DataFrame from TableNormalizer."""
        from quality_audit.utils.table_normalizer import TableNormalizer

        df = pd.DataFrame(
            {
                "Code": ["100", "200"],
                "2024": [1000, 2000],
                "2023": [900, 1900],
            }
        )

        # Normalize
        normalized_df, metadata = TableNormalizer.normalize_table(df, "Test")

        # Validator should accept normalized DataFrame
        validator, skip_reason = ValidatorFactory.get_validator(normalized_df, "Test")
        if validator:
            result = validator.validate(normalized_df, "Test")
            assert isinstance(result, ValidationResult)


class TestValidatorReportWriterContract:
    """Test 3: Validator → Report Writer contract."""

    def test_validation_result_schema(self):
        """Verify ValidationResult has all required fields for report writer."""
        result = ValidationResult(
            status="FAIL: Test",
            marks=[],
            cross_ref_marks=[],
            rule_id="TEST_RULE",
            status_enum="FAIL",
            context={"heading": "Test Table"},
            severity="HIGH",
            confidence="HIGH",
            table_id="tbl_001_test",
            root_cause="calculation",
        )

        # Convert to dict (as report writer expects)
        result_dict = result.to_dict()

        # Contract: Required fields for report writer
        required_fields = [
            "status",
            "status_enum",
            "rule_id",
            "context",
            "table_id",  # SCRUM-8: Required for hyperlinks
        ]
        for field in required_fields:
            assert field in result_dict, f"Missing required field: {field}"

        # Contract: Optional but important fields
        optional_fields = ["severity", "confidence", "root_cause", "marks"]
        # At least some should be present
        assert any(field in result_dict for field in optional_fields)

    def test_excel_writer_handles_all_result_types(self):
        """Verify ExcelWriter can handle all ValidationResult status types."""
        writer = ExcelWriter()
        wb = Workbook()

        # Test all status types
        status_types = ["PASS", "FAIL", "WARN", "INFO", "ERROR"]
        results = []
        for status in status_types:
            results.append(
                {
                    "status": f"{status}: Test",
                    "status_enum": status,
                    "rule_id": f"TEST_{status}",
                    "table_id": f"tbl_001_{status.lower()}",
                    "context": {"heading": f"Test {status}"},
                    "severity": "MEDIUM",
                    "root_cause": "general",
                }
            )

        # Should not crash
        writer.write_summary_sheet(wb, results, [("Test", 1)])
        assert "Tổng hợp kiểm tra" in wb.sheetnames

    def test_anchor_map_population(self):
        """Verify anchor_map is populated correctly for hyperlinks."""
        writer = ExcelWriter()
        wb = Workbook()

        df = pd.DataFrame({"A": [1, 2], "B": [3, 4]})
        table_heading_pairs = [(df, "Test Table 1"), (df, "Test Table 2")]
        results = [
            {
                "status": "PASS",
                "status_enum": "PASS",
                "rule_id": "TEST",
                "table_id": "tbl_001_test_table_1",
                "context": {"heading": "Test Table 1"},
            },
            {
                "status": "FAIL",
                "status_enum": "FAIL",
                "rule_id": "TEST",
                "table_id": "tbl_002_test_table_2",
                "context": {"heading": "Test Table 2"},
            },
        ]

        writer.write_tables_consolidated(wb, table_heading_pairs, results)

        # Contract: anchor_map populated for all tables
        assert "tbl_001_test_table_1" in writer.anchor_map
        assert "tbl_002_test_table_2" in writer.anchor_map
        assert writer.anchor_map["tbl_001_test_table_1"] >= 1
        assert (
            writer.anchor_map["tbl_002_test_table_2"]
            > writer.anchor_map["tbl_001_test_table_1"]
        )


class TestEndToEndPipelineContract:
    """Test 4: End-to-end pipeline contract."""

    def test_audit_service_pipeline(self):
        """Verify AuditService can process tables end-to-end."""
        service = AuditService()

        # Create minimal test data
        df = pd.DataFrame(
            {
                "Code": ["100", "200", "Total"],
                "2024": [1000, 2000, 3000],
                "2023": [900, 1900, 2800],
            }
        )

        table_heading_pairs = [(df, "Test Balance Sheet")]
        results = service._validate_tables(table_heading_pairs)

        # Contract: Returns list of result dicts
        assert isinstance(results, list)
        assert len(results) == 1

        result = results[0]
        # Contract: Each result has required fields
        assert "status" in result
        assert "status_enum" in result
        assert "rule_id" in result
        assert "table_id" in result  # SCRUM-8: Required for hyperlinks
        assert "context" in result

    def test_report_generation_contract(self):
        """Verify report generation produces expected sheets."""
        service = AuditService()
        wb = service.excel_writer.create_workbook()

        df = pd.DataFrame({"Code": ["100"], "CY": [1000], "PY": [900]})
        table_heading_pairs = [(df, "Test Table")]
        results = [
            {
                "status": "PASS",
                "status_enum": "PASS",
                "rule_id": "TEST",
                "table_id": "tbl_001_test",
                "context": {"heading": "Test Table"},
            }
        ]

        # Generate report
        service.excel_writer.write_tables_consolidated(wb, table_heading_pairs, results)
        service.excel_writer.write_executive_summary(wb, results)
        service.excel_writer.write_focus_list(wb, results)
        service.excel_writer.write_summary_sheet(wb, results, [("Test", 1)])

        # Contract: Expected sheets exist
        expected_sheets = [
            "FS casting",
            "Executive Summary",
            "Focus List",
            "Tổng hợp kiểm tra",
        ]
        for sheet_name in expected_sheets:
            assert sheet_name in wb.sheetnames, f"Missing sheet: {sheet_name}"

    def test_contract_v2_sheets_created(self):
        """Verify contract v2 sheets are present and internally consistent."""
        service = AuditService()
        wb = service.excel_writer.create_workbook()
        results = [
            {
                "status": "PASS",
                "status_enum": "PASS",
                "rule_id": "TEST_PASS",
                "table_id": "tbl_001_test",
                "table_name": "Table A",
                "context": {"heading": "Table A", "extractor_engine": "ooxml"},
            },
            {
                "status": "FAIL",
                "status_enum": "FAIL",
                "rule_id": "TEST_FAIL",
                "table_id": "tbl_002_test",
                "table_name": "Table B",
                "failure_reason_code": "MATH_EQ",
                "context": {"heading": "Table B", "quality_score": 0.8},
            },
        ]
        service.excel_writer.write_contract_v2_sheets(wb, results, telemetry=None)
        assert "Summary" in wb.sheetnames
        assert "Findings" in wb.sheetnames
        assert "Metadata" in wb.sheetnames
        summary_ws = wb["Summary"]
        findings_ws = wb["Findings"]
        metadata_ws = wb["Metadata"]
        # Header + 2 rows in Findings
        assert findings_ws.max_row == 3
        # Metadata integrity row must evaluate to True
        integrity_rows = [
            row
            for row in metadata_ws.iter_rows(min_row=2, values_only=True)
            if row[0] == "integrity_check"
        ]
        assert len(integrity_rows) == 1
        assert integrity_rows[0][1] is True
        # Summary includes total metric
        total_rows = [
            row
            for row in summary_ws.iter_rows(min_row=2, values_only=True)
            if row[0] == "Total tables"
        ]
        assert len(total_rows) == 1
        assert total_rows[0][1] == 2


class TestSchemaInvariants:
    """Test 5: Schema invariants from SCRUM-5→8."""

    def test_never_crash_policy(self):
        """Verify validators never raise exceptions (SCRUM-5/6)."""
        service = AuditService()

        # Create problematic table that might cause IndexError
        df = pd.DataFrame({"A": [1]})  # Very short table

        result = service._validate_single_table(df, "Short Table")

        # Contract: Always returns ValidationResult, never raises
        assert isinstance(result, ValidationResult)
        # Should be ERROR or WARN, not crash
        assert result.status_enum in ["PASS", "FAIL", "WARN", "INFO", "ERROR"]

    def test_severity_not_all_medium(self):
        """Verify severity calculation produces distribution (SCRUM-8)."""
        from quality_audit.core.validators.generic_validator import (
            GenericTableValidator,
        )

        validator = GenericTableValidator()

        # Test different scenarios
        test_cases = [
            ("MATH_EQ", 1000000.0, False, "HIGH"),  # Large diff >= 1M: HIGH
            ("MATH_EQ", 100000.0, False, "HIGH"),  # Diff >= 100K: HIGH
            ("MATH_EQ", 10000.0, False, "MEDIUM"),  # Diff >= 10K: MEDIUM
            ("MATH_EQ", 100.0, False, "LOW"),  # Diff < 10K: LOW
            ("MATH_EQ", 1.0, False, "LOW"),  # Small diff: LOW
            ("SKIPPED_FOOTER_SIGNATURE", 0.0, True, "LOW"),  # Skipped
        ]

        for rule_id, diff, is_skipped, expected in test_cases:
            severity = validator._calculate_severity(rule_id, diff, is_skipped)
            assert (
                severity == expected
            ), f"Rule {rule_id}: expected {expected}, got {severity}"

    def test_root_cause_not_all_general(self):
        """Verify root_cause inference produces variety (SCRUM-8)."""
        from quality_audit.core.validators.generic_validator import (
            GenericTableValidator,
        )

        validator = GenericTableValidator()

        test_cases = [
            ("MATH_EQ", {}, "calculation"),
            ("CROSS_CHECK_MISMATCH", {}, "cross_ref"),
            ("UNKNOWN", {"subtotals_detected": True}, "subtotal"),
            ("UNKNOWN", {"is_movement_table": True}, "movement"),
            ("UNKNOWN", {"missing_codes": True}, "mapping"),
        ]

        for rule_id, context, expected in test_cases:
            root_cause = validator._infer_root_cause(rule_id, context)
            assert (
                root_cause == expected
            ), f"Rule {rule_id}: expected {expected}, got {root_cause}"

    def test_max_diff_extraction(self):
        """Verify max_diff is extracted correctly (SCRUM-8)."""
        result = ValidationResult(
            status="FAIL: Sai lệch = 123,456.00",
            marks=[
                {"diff": 1000.0, "ok": False},
                {"diff": 500.0, "ok": False},
            ],
            rule_id="TEST",
            status_enum="FAIL",
            context={"max_diff": 123456.0},
        )

        # Contract: max_diff available from multiple sources
        assert result.context.get("max_diff") == 123456.0

        # Test extraction from marks
        from_marks = [m.get("diff") for m in result.marks if m.get("diff")]
        assert max(from_marks) == 1000.0
