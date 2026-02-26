"""
Script to extract FAIL_TOOL_EXTRACT table IDs from audit XLSX output.
Creates gold set manifest for P1-1 implementation.

Usage (from repo root):
  python scripts/extract_gold_set.py [--results-dir DIR] [--manifest PATH]
  Default: --results-dir=./results, --manifest=tests/fixtures/gold_set_manifest.json
"""

import argparse
import io
import json
import sys

# Fix Windows console encoding
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

from pathlib import Path  # noqa: E402

import openpyxl  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract FAIL_TOOL_EXTRACT table IDs from audit XLSX output."
    )
    parser.add_argument(
        "--results-dir",
        type=Path,
        default=Path("results"),
        help="Directory containing *_output.xlsx files (default: ./results)",
    )
    parser.add_argument(
        "--manifest",
        type=Path,
        default=Path("tests/fixtures/gold_set_manifest.json"),
        help="Output manifest path (default: tests/fixtures/gold_set_manifest.json)",
    )
    return parser.parse_args()


args = parse_args()
results_dir = args.results_dir.resolve()
manifest_path = args.manifest.resolve()

fail_tables = []

for xlsx_file in results_dir.glob("*_output.xlsx"):
    print(f"\n=== {xlsx_file.name} ===")
    wb = openpyxl.load_workbook(xlsx_file, data_only=True)

    # Find the summary sheet (might be Vietnamese name)
    print(f"Available sheets: {wb.sheetnames}")
    summary_sheet = None
    for name in wb.sheetnames:
        # Look for Vietnamese keywords
        if "hợp" in name or "kiểm" in name or "tong hop" in name.lower():
            summary_sheet = wb[name]
            break

    # Fallback: if first sheet is Executive Summary, use second
    if not summary_sheet:
        if len(wb.sheetnames) > 1 and "executive" in wb.sheetnames[0].lower():
            summary_sheet = wb[wb.sheetnames[1]]
        else:
            summary_sheet = wb[wb.sheetnames[0]]

    print(f"Using sheet: {summary_sheet.title}")

    # Find headers
    headers = {}
    for col in range(1, 30):
        val = summary_sheet.cell(1, col).value
        if val:
            headers[col] = str(val)

    # Find key columns
    status_col = None
    table_id_col = None
    file_col = None
    table_idx_col = None

    for col, header in headers.items():
        h = header.lower()
        if "status" in h or "enum" in h:
            status_col = col
        if "table" in h and "id" in h:
            table_id_col = col
        if "file" in h or "nguon" in h:
            file_col = col
        if "index" in h or "stt" in h:
            table_idx_col = col

    print(f"Status col: {status_col}, TableID col: {table_id_col}")

    # Extract FAIL_TOOL_EXTRACT rows
    for row in range(2, summary_sheet.max_row + 1):
        status = summary_sheet.cell(row, status_col).value if status_col else None
        if status and "FAIL_TOOL_EXTRACT" in str(status):
            table_id = (
                summary_sheet.cell(row, table_id_col).value
                if table_id_col
                else f"row_{row}"
            )
            file_name = (
                summary_sheet.cell(row, file_col).value if file_col else xlsx_file.stem
            )
            table_idx = (
                summary_sheet.cell(row, table_idx_col).value
                if table_idx_col
                else row - 1
            )

            fail_tables.append(
                {
                    "source_file": str(xlsx_file.name),
                    "table_id": str(table_id),
                    "file_name": str(file_name) if file_name else "unknown",
                    "table_index": int(table_idx) if table_idx else row - 1,
                    "status": str(status),
                }
            )

print(f"\n=== FAIL_TOOL_EXTRACT Tables: {len(fail_tables)} ===")
for t in fail_tables:
    print(f"  - {t['table_id']} (idx {t['table_index']}) from {t['file_name']}")

# Save manifest
manifest_path.parent.mkdir(parents=True, exist_ok=True)

manifest = {
    "version": 1,
    "description": "P1-1 Gold Set: FAIL_TOOL_EXTRACT tables for regression testing",
    "fail_tool_extract_tables": fail_tables,
    "expected_count": len(fail_tables),
    "edge_cases": {
        "rule_b_blank_label": [],
        "equity_validator": [],
        "tax_validator": [],
    },
}

with open(manifest_path, "w", encoding="utf-8") as f:
    json.dump(manifest, f, indent=2, ensure_ascii=False)

print(f"\nManifest saved to: {manifest_path}")
