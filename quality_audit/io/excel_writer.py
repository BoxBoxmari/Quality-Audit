"""
Excel workbook creation and writing utilities with security sanitization.
"""

import contextlib
import re
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..config.constants import GREEN_FILL, INFO_FILL, RED_FILL
from ..config.feature_flags import get_feature_flags
from ..utils.formatters import (
    apply_cell_marks,
    apply_crossref_marks,
    sanitize_excel_value,
)
from ..utils.table_canonicalizer import TableMeta, canonicalize_table

# Font attributes supported by openpyxl Font (and StyleProxy)
_FONT_ATTRS = (
    "name",
    "size",
    "bold",
    "italic",
    "underline",
    "strike",
    "color",
    "outline",
    "shadow",
    "condense",
    "extend",
    "vertAlign",
    "charset",
    "family",
    "scheme",
)


def _font_with(font, **overrides):
    """Return a new Font with same attributes as font plus overrides. Replaces deprecated font.copy(**kw)."""
    kw = {}
    for attr in _FONT_ATTRS:
        try:
            val = getattr(font, attr, None)
            if val is not None:
                kw[attr] = val
        except Exception:
            pass
    kw.update(overrides)
    return Font(**kw)


class ExcelWriter:
    """Handles secure Excel workbook creation and writing."""

    def __init__(self, previous_output_path: Optional[str] = None):
        """Initialize Excel writer.

        Args:
            previous_output_path: Optional path to previous Excel output for triage carry-forward
        """
        self.previous_output_path = previous_output_path
        # SCRUM-8: Anchor map for direct hyperlinks (table_id -> row number in FS casting)
        self.anchor_map: Dict[str, int] = {}
        # SCRUM-8: Previous triage data for carry-forward
        self._previous_triage_data: Dict[str, Dict[str, str]] = {}

    def create_workbook(self) -> Workbook:
        """
        Create a new Excel workbook.

        Returns:
            Workbook: New OpenPyXL workbook
        """
        return Workbook()

    def write_summary_sheet(
        self,
        wb: Workbook,
        results: List[Dict],
        sheet_positions: List[Tuple[str, int]],
        telemetry=None,
    ) -> None:
        """
        Write summary sheet with validation results and hyperlinks.

        SCRUM-8: Hyperlinks point to FS casting sheet using anchor_map.
        SCRUM-8: Conditional formatting only on Status Enum column (column C).
        Phase 6: Extractor Engine, Quality Score, Failure Reason Code, run_id.

        Args:
            wb: Excel workbook
            results: List of validation results
            sheet_positions: List of (heading, start_row) tuples
            telemetry: Optional TelemetryCollector for run_id
        """
        # Create new sheet instead of using wb.active to avoid overwriting Executive Summary
        if "Tổng hợp kiểm tra" in wb.sheetnames:
            ws = wb["Tổng hợp kiểm tra"]
            ws.delete_rows(1, ws.max_row)  # Clear existing content
        else:
            ws = wb.create_sheet(title="Tổng hợp kiểm tra")
        # Ensure this sheet is active for downstream expectations/tests
        from contextlib import suppress

        with suppress(Exception):
            wb.active = wb.sheetnames.index("Tổng hợp kiểm tra")

        run_id = ""
        if telemetry is not None and getattr(telemetry, "run_telemetry", None):
            run_id = getattr(telemetry.run_telemetry, "run_id", "") or ""

        # Headers:
        # A: Tên bảng, B: Trạng thái kiểm tra, C: Status Enum, D: Status Category (WARN taxonomy), E: Rule ID, F: Validator Type
        # Phase 6: G: Extractor Engine, H: Quality Score, I: Failure Reason Code, J: run_id
        # R1: K–P per-table baseline; P1 forensic: Q–T; P0: U–AA metadata (heading, classifier, extractor usable, assertions)
        ws.append(
            [
                "Tên bảng",
                "Trạng thái kiểm tra",
                "Status Enum",
                "Status Category",
                "Rule ID",
                "Validator Type",
                "Extractor Engine",
                "Quality Score",
                "Failure Reason Code",
                "run_id",
                "Engine Attempts",
                "Invariants Failed",
                "Grid Cols Expected",
                "Grid Cols Built",
                "GridSpan Count",
                "vMerge Count",
                "Total Row Method",
                "Total Row Index",
                "Total Row Confidence",
                "Column Classification Method",
                "Render First Rejection",
                "Render First Confidence",
                "Render First Coverage",
                "Excluded Columns",
                "Heading Source",
                "Heading Confidence",
                "Classifier Type",
                "Classifier Confidence",
                "Classifier Reason",
                "Extractor Usable Reason",
                "Assertions Count",
                "Numeric Evidence",
                "CY Column",
                "PY Column",
                "Reason Code",
            ]
        )

        for i, (result, (heading, _start_row)) in enumerate(
            zip(results, sheet_positions)
        ):
            row_idx = i + 2
            ctx = result.get("context") or {}
            total_row_meta = ctx.get("total_row_metadata") or {}
            # Column A: Table Name with hyperlink to FS casting (sanitize for formula injection)
            table_id = result.get("table_id")
            cell = ws.cell(
                row=row_idx, column=1, value=sanitize_excel_value(heading or "")
            )

            # SCRUM-8: Hyperlink to FS casting using anchor_map (not individual sheets)
            if table_id and table_id in self.anchor_map:
                anchor_row = self.anchor_map[table_id]
                cell.hyperlink = f"#'FS casting'!A{anchor_row}"
                cell.style = "Hyperlink"
            elif table_id:
                # Anchor missing - diagnostic flag
                cell.value = f"{heading} (missing anchor)"
                result.setdefault("context", {})["anchor_missing"] = True

            # Column B: Status message (sanitize for formula injection)
            ws.cell(
                row=row_idx,
                column=2,
                value=sanitize_excel_value(result.get("status") or ""),
            )

            # Column C: Status Enum (for conditional formatting)
            status_enum = result.get("status_enum", "UNKNOWN")
            ws.cell(row=row_idx, column=3, value=status_enum)

            # Column D: Status Category (sanitize for formula injection)
            ws.cell(
                row=row_idx,
                column=4,
                value=sanitize_excel_value(result.get("status_category") or ""),
            )

            # Column E: Rule ID (sanitize for formula injection)
            ws.cell(
                row=row_idx,
                column=5,
                value=sanitize_excel_value(result.get("rule_id") or "UNKNOWN"),
            )

            # Column F: Validator Type (sanitize for formula injection)
            validator_type = sanitize_excel_value(ctx.get("validator_type") or "")
            v_cell = ws.cell(row=row_idx, column=6, value=validator_type)
            v_cell.font = Font(color="00000000", underline=None)

            # Phase 6: G, H, I, J (sanitize text for formula injection)
            ws.cell(
                row=row_idx,
                column=7,
                value=sanitize_excel_value(ctx.get("extractor_engine") or ""),
            )
            qs = ctx.get("quality_score")
            ws.cell(
                row=row_idx, column=8, value=qs if qs is None else round(float(qs), 4)
            )
            ws.cell(
                row=row_idx,
                column=9,
                value=sanitize_excel_value(
                    result.get("failure_reason_code")
                    or ctx.get("failure_reason_code")
                    or ""
                ),
            )
            ws.cell(row=row_idx, column=10, value=run_id)
            # R1: K–P per-table baseline
            engine_attempts = ctx.get("engine_attempts")
            ws.cell(
                row=row_idx,
                column=11,
                value=(
                    ",".join(engine_attempts)
                    if isinstance(engine_attempts, list)
                    else engine_attempts
                ),
            )
            invariants_failed = ctx.get("invariants_failed")
            ws.cell(
                row=row_idx,
                column=12,
                value=(
                    ",".join(invariants_failed)
                    if isinstance(invariants_failed, list)
                    else invariants_failed
                ),
            )
            ws.cell(row=row_idx, column=13, value=ctx.get("grid_cols_expected"))
            ws.cell(row=row_idx, column=14, value=ctx.get("grid_cols_built"))
            ws.cell(row=row_idx, column=15, value=ctx.get("gridSpan_count"))
            ws.cell(row=row_idx, column=16, value=ctx.get("vMerge_count"))
            # P1 forensic: Q–T from total_row_metadata and context
            ws.cell(row=row_idx, column=17, value=total_row_meta.get("method"))
            ws.cell(row=row_idx, column=18, value=total_row_meta.get("total_row_idx"))
            ws.cell(row=row_idx, column=19, value=total_row_meta.get("confidence"))
            _ccm = ctx.get("column_classification_method")
            ws.cell(
                row=row_idx,
                column=20,
                value=sanitize_excel_value(_ccm) if isinstance(_ccm, str) else _ccm,
            )
            # P0-1/P0-3: Render-first telemetry and excluded_columns
            render_first_meta = ctx.get("render_first_metadata") or {}
            ws.cell(
                row=row_idx,
                column=21,
                value=sanitize_excel_value(
                    render_first_meta.get("rejection_reason") or ""
                ),
            )
            rf_conf = render_first_meta.get("mean_cell_confidence")
            ws.cell(
                row=row_idx,
                column=22,
                value=round(float(rf_conf), 4) if rf_conf is not None else None,
            )
            rf_cov = render_first_meta.get("token_coverage_ratio")
            ws.cell(
                row=row_idx,
                column=23,
                value=round(float(rf_cov), 4) if rf_cov is not None else None,
            )
            excluded_cols = ctx.get("excluded_columns")
            ws.cell(
                row=row_idx,
                column=24,
                value=(
                    ",".join(str(c) for c in excluded_cols)
                    if isinstance(excluded_cols, (list, tuple))
                    else sanitize_excel_value(excluded_cols or "")
                ),
            )
            # Phase 0: Heading / classifier / extractor metadata and assertions_count
            ws.cell(
                row=row_idx,
                column=25,
                value=sanitize_excel_value(ctx.get("heading_source") or ""),
            )
            heading_conf = ctx.get("heading_confidence")
            ws.cell(
                row=row_idx,
                column=26,
                value=(
                    round(float(heading_conf), 4) if heading_conf is not None else None
                ),
            )
            ws.cell(
                row=row_idx,
                column=27,
                value=sanitize_excel_value(ctx.get("classifier_primary_type") or ""),
            )
            cl_conf = ctx.get("classifier_confidence")
            ws.cell(
                row=row_idx,
                column=28,
                value=round(float(cl_conf), 4) if cl_conf is not None else None,
            )
            ws.cell(
                row=row_idx,
                column=29,
                value=sanitize_excel_value(ctx.get("classifier_reason") or ""),
            )
            ws.cell(
                row=row_idx,
                column=30,
                value=sanitize_excel_value(ctx.get("extractor_usable_reason") or ""),
            )
            ws.cell(
                row=row_idx,
                column=31,
                value=result.get("assertions_count", ctx.get("assertions_count")),
            )
            # P4.2: Diagnostic columns - numeric evidence, CY/PY, reason_code
            num_ev = ctx.get("numeric_evidence_score")
            ws.cell(
                row=row_idx,
                column=32,
                value=round(float(num_ev), 4) if num_ev is not None else None,
            )
            cy_col = ctx.get("cy_column")
            if cy_col is None and isinstance(ctx.get("amount_columns"), (list, tuple)):
                ac = ctx.get("amount_columns")
                cy_col = ac[0] if ac else None
            ws.cell(
                row=row_idx,
                column=33,
                value=sanitize_excel_value(cy_col if cy_col is not None else ""),
            )
            py_col = ctx.get("py_column")
            ac = ctx.get("amount_columns")
            if py_col is None and isinstance(ac, (list, tuple)) and len(ac) > 1:
                py_col = ac[1]
            ws.cell(
                row=row_idx,
                column=34,
                value=sanitize_excel_value(py_col if py_col is not None else ""),
            )
            reason_code = (
                ctx.get("reason_code")
                or result.get("failure_reason_code")
                or ctx.get("failure_reason_code")
                or ""
            )
            ws.cell(
                row=row_idx,
                column=35,
                value=sanitize_excel_value(reason_code),
            )

        # SCRUM-8: Apply conditional formatting ONLY to Status Enum column (column C)
        self._apply_status_colors_to_enum_column(ws)

    def write_tables(
        self,
        wb: Workbook,
        table_heading_pairs: List[Tuple[pd.DataFrame, str]],
        results: List[Dict],
    ) -> List[str]:
        """
        Write all tables to individual sheets.

        Args:
            wb: Excel workbook
            table_heading_pairs: List of (table_df, heading) tuples
            results: List of validation results

        Returns:
            List[str]: List of created sheet names
        """
        sheet_names = []

        for i, ((table, heading), result) in enumerate(
            zip(table_heading_pairs, results)
        ):
            # Sanitize table data
            table = table.copy()
            table = table.map(sanitize_excel_value)

            # Create sheet name
            flags = get_feature_flags()
            note_num = result.get("context", {}).get("note_number")
            raw_name = heading if heading else f"Bảng {i + 1}"
            if note_num and flags.get("ENABLE_NOTE_NUMBER_MAPPING", True):
                raw_name = f"[Note {note_num}] {raw_name}"

            # Excel limits sheet names to 31 characters.
            # We reserve 4-5 chars for suffixes (e.g. "_999"), so max base length is 26.
            sheet_name = self._shorten_sheet_name(raw_name, max_length=26)

            # Ensure unique sheet name
            original_name = sheet_name
            counter = 2
            while sheet_name in wb.sheetnames:
                suffix = f"_{counter}"
                # Safely truncate original name if suffix makes it > 31
                cutoff = 31 - len(suffix)
                sheet_name = f"{original_name[:cutoff]}{suffix}"
                counter += 1

            sheet_names.append(sheet_name)

            # Create worksheet and write data
            ws = wb.create_sheet(title=sheet_name)

            # Write DataFrame to worksheet
            for row_idx, row in enumerate(
                dataframe_to_rows(table, index=False, header=True)
            ):
                for col_idx, value in enumerate(row):
                    # Format numbers as accounting format
                    if row_idx > 0:  # Skip header row
                        try:
                            num_val = float(
                                str(value)
                                .replace(",", "")
                                .replace("(", "-")
                                .replace(")", "")
                            )
                            cell = ws.cell(
                                row=row_idx + 1, column=col_idx + 1, value=num_val
                            )
                            cell.number_format = (
                                '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                            )
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                    else:
                        ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)

            # Create Excel table
            max_col = get_column_letter(ws.max_column)
            max_row = ws.max_row
            tab = Table(displayName=f"Table{i + 1}", ref=f"A1:{max_col}{max_row}")
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9", showRowStripes=True
            )
            ws.add_table(tab)

            # Apply formatting
            apply_cell_marks(ws, result.get("marks", []))
            apply_crossref_marks(ws, result.get("cross_ref_marks", []))

            # Add status and back link
            ws.cell(row=max_row + 2, column=1, value="Trạng thái kiểm tra:")
            ws.cell(row=max_row + 2, column=2, value=result.get("status"))

            back_cell = ws.cell(row=max_row + 3, column=1, value="⬅ Quay lại Tổng hợp")
            back_cell.hyperlink = "#'Tổng hợp kiểm tra'!A1"
            back_cell.style = "Hyperlink"

        return sheet_names

    def write_tables_consolidated(
        self,
        wb: Workbook,
        table_heading_pairs: List[Tuple[pd.DataFrame, Optional[str]]],
        results: List[Dict],
    ) -> List[Tuple[str, int]]:
        """
        Write all tables to a single consolidated sheet.

        SCRUM-8: Populates anchor_map for direct hyperlinks.

        Args:
            wb: Excel workbook
            table_heading_pairs: List of (table_df, heading) tuples
            results: List of validation results

        Returns:
            List[Tuple[str, int]]: List of (heading, start_row) tuples
        """
        # Reset anchor_map for new workbook
        self.anchor_map = {}

        sheet_name = "FS casting"
        ws = wb.create_sheet(title=sheet_name)
        sheet_positions = []
        current_row = 1

        flags = get_feature_flags()
        enable_canonicalize_writer = flags.get("enable_canonicalize_writer", False)

        for i, ((table, heading), result) in enumerate(
            zip(table_heading_pairs, results)
        ):
            # Sanitize table data
            table = table.copy()
            table.columns = table.columns.map(str)

            # Add heading with visible table_id (SCRUM-8)
            note_num = result.get("context", {}).get("note_number")
            raw_name = heading if heading else f"Bảng {i + 1}"
            if note_num:
                raw_name = f"[Note {note_num}] {raw_name}"
            sheet_positions.append((raw_name, current_row))

            # SCRUM-8: Get table_id from result and add to anchor_map
            # Fallback: Generate table_id if missing (should not happen, but ensures robustness)
            table_id = result.get("table_id")
            if not table_id:
                # Generate deterministic table_id matching audit_service format
                safe_heading = re.sub(r"[^A-Za-z0-9]", "_", heading or "unknown")
                slug = safe_heading[:50].strip("_")
                if not slug:
                    slug = "unnamed"
                table_id = f"tbl_{i + 1:03d}_{slug}"
                table_id = re.sub(r"[^A-Za-z0-9_]", "_", table_id)

            # Canonicalize before writing when flag is on (reduce column explosion, index-row, Code.*)
            if enable_canonicalize_writer:
                table_meta = TableMeta(table_id=table_id, table_no=i + 1)
                table, _canon_report = canonicalize_table(table, table_meta)

            # Always populate anchor_map for every table (SCRUM-8: ensures all hyperlinks work)
            self.anchor_map[table_id] = current_row
            # Create named range for direct hyperlink
            quoted_sheet = quote_sheetname(sheet_name)
            ref = f"'{quoted_sheet}'!$A${current_row}"
            wb.defined_names[table_id] = DefinedName(table_id, attr_text=ref)
            # Header with visible ID
            header_value = f"{raw_name} [{table_id}]"

            ws.cell(row=current_row, column=1, value=header_value)
            current_row += 1

            # Write table
            start_row = current_row
            start_col = 1

            for row_idx, row in enumerate(
                dataframe_to_rows(table, index=False, header=True)
            ):
                for col_idx, value in enumerate(row):
                    if row_idx == 0:
                        # Header row - no sanitization needed for headers
                        ws.cell(
                            row=current_row, column=start_col + col_idx, value=value
                        )
                    else:
                        # Data row - try numeric formatting first, then sanitize if
                        # needed
                        try:
                            # Try to parse as number first
                            num_val = float(
                                str(value)
                                .replace(",", "")
                                .replace("(", "-")
                                .replace(")", "")
                            )
                            cell = ws.cell(
                                row=current_row,
                                column=start_col + col_idx,
                                value=num_val,
                            )
                            cell.number_format = (
                                '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
                            )
                        except (ValueError, TypeError):
                            # Not a number, sanitize and write as text
                            safe_value = sanitize_excel_value(value)
                            ws.cell(
                                row=current_row,
                                column=start_col + col_idx,
                                value=safe_value,
                            )

                current_row += 1

            # Create Excel table
            end_row = current_row - 1
            end_col = start_col + len(table.columns) - 1
            table_range = f"{ws.cell(row=start_row, column=start_col).coordinate}:{ws.cell(row=end_row, column=end_col).coordinate}"
            excel_table = Table(displayName=f"Table_{i + 1}", ref=table_range)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            excel_table.tableStyleInfo = style
            ws.add_table(excel_table)

            # Apply formatting
            apply_cell_marks(ws, result.get("marks", []), start_row - 1, start_col - 1)
            apply_crossref_marks(
                ws, result.get("cross_ref_marks", []), start_row, start_col - 1
            )

            # Add status and spacing
            ws.cell(row=current_row + 1, column=1, value="Trạng thái kiểm tra:")
            ws.cell(row=current_row + 1, column=2, value=result.get("status"))

            back_cell = ws.cell(
                row=current_row + 2, column=1, value="⬅ Quay lại Tổng hợp"
            )
            back_cell.hyperlink = "#'Tổng hợp kiểm tra'!A1"
            back_cell.style = "Hyperlink"

            current_row += 4  # Add spacing between tables

        return sheet_positions

    def save_workbook(self, wb: Workbook, file_path: str) -> None:
        """
        Save workbook to file.

        Args:
            wb: Excel workbook
            file_path: Output file path
        """
        wb.save(file_path)

    def _shorten_sheet_name(self, name: str, max_length: int = 20) -> str:
        """
        Shorten sheet name for Excel compatibility.

        Args:
            name: Original name
            max_length: Maximum length

        Returns:
            str: Shortened name
        """
        from ..config.constants import _SHEET_NAME_CLEAN_RE

        name = _SHEET_NAME_CLEAN_RE.sub("_", str(name))
        return name[:max_length]

    def _apply_status_colors(self, ws) -> None:
        """
        Apply color coding to status cells.

        Args:
            ws: Worksheet to format
        """
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                status_text = str(cell.value or "").lower()
                if "PASS" in status_text:
                    cell.fill = GREEN_FILL
                elif "FAIL" in status_text:
                    cell.fill = RED_FILL
                elif "WARN" in status_text:
                    cell.fill = INFO_FILL
                else:
                    cell.fill = INFO_FILL  # Default for unknown status

    def _apply_status_colors_to_enum_column(self, ws) -> None:
        """
        SCRUM-8: Apply conditional formatting ONLY to Status Enum column (column C).

        Uses Status Enum values (PASS/FAIL/WARN/INFO/ERROR) for deterministic formatting.
        Does not format column B (message text) to avoid unintended highlighting.

        Args:
            ws: Worksheet to format
        """
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import Font, PatternFill

        # Remove any existing conditional formatting on column C
        if ws.conditional_formatting:
            # Clear existing rules for column C
            ws.conditional_formatting = {
                ref: rules
                for ref, rules in ws.conditional_formatting.items()
                if not ref.startswith("C")
            }

        # Define color fills
        pass_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )
        fail_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        warn_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )  # Amber
        info_fill = PatternFill(
            start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"
        )  # Grey/Blue
        error_fill = PatternFill(
            start_color="FF0000", end_color="FF0000", fill_type="solid"
        )  # Red
        error_font = Font(color="FFFFFF", bold=True)  # White bold for ERROR

        # Apply conditional formatting rules to column C (Status Enum column)
        # Use stopIfTrue=True to prevent overlap
        max_row = ws.max_row
        if max_row > 1:  # Only if there are data rows
            col_range = f"C2:C{max_row}"

            # Build list of rules (order matters - highest priority first)
            rules = []

            # Rule 1: ERROR - Red fill with white bold text (highest priority)
            error_rule = CellIsRule(
                operator="equal",
                formula=["ERROR"],
                fill=error_fill,
                font=error_font,
                stopIfTrue=True,
            )
            rules.append(error_rule)

            # Rule 2: FAIL_TOOL_EXTRACT - Red fill
            fail_tool_extract_rule = CellIsRule(
                operator="equal",
                formula=["FAIL_TOOL_EXTRACT"],
                fill=fail_fill,
                stopIfTrue=True,
            )
            rules.append(fail_tool_extract_rule)

            # Rule 3: FAIL_TOOL_LOGIC - Red fill
            fail_tool_logic_rule = CellIsRule(
                operator="equal",
                formula=["FAIL_TOOL_LOGIC"],
                fill=fail_fill,
                stopIfTrue=True,
            )
            rules.append(fail_tool_logic_rule)

            # Rule 4: FAIL_DATA - Red fill
            fail_data_rule = CellIsRule(
                operator="equal",
                formula=["FAIL_DATA"],
                fill=fail_fill,
                stopIfTrue=True,
            )
            rules.append(fail_data_rule)

            # Rule 5: FAIL - Red fill
            fail_rule = CellIsRule(
                operator="equal",
                formula=["FAIL"],
                fill=fail_fill,
                stopIfTrue=True,
            )
            rules.append(fail_rule)

            # Rule 6: WARN - Amber fill
            warn_rule = CellIsRule(
                operator="equal",
                formula=["WARN"],
                fill=warn_fill,
                stopIfTrue=True,
            )
            rules.append(warn_rule)

            # Rule 7: INFO - Grey/Blue fill
            info_rule = CellIsRule(
                operator="equal",
                formula=["INFO"],
                fill=info_fill,
                stopIfTrue=True,
            )
            rules.append(info_rule)

            # Rule 8: PASS - Green fill (optional, lowest priority)
            pass_rule = CellIsRule(
                operator="equal",
                formula=["PASS"],
                fill=pass_fill,
                stopIfTrue=True,
            )
            rules.append(pass_rule)

            # Apply all rules to the range using .add() method (correct openpyxl API)
            # Each rule must be added individually
            for rule in rules:
                ws.conditional_formatting.add(col_range, rule)

            # Mirror fills directly onto column C cells.
            # Some tests/consumers inspect `cell.fill` (not conditional formatting rules),
            # so we apply deterministic fills to the cells themselves.
            for row in ws.iter_rows(min_row=2, min_col=3, max_col=3, max_row=max_row):
                for cell in row:
                    status_enum = str(cell.value or "").strip().upper()
                    if status_enum == "ERROR":
                        cell.fill = error_fill
                        cell.font = error_font
                    elif status_enum in (
                        "FAIL_TOOL_EXTRACT",
                        "FAIL_TOOL_LOGIC",
                        "FAIL_DATA",
                        "FAIL",
                    ):
                        cell.fill = fail_fill
                    elif status_enum == "WARN":
                        cell.fill = warn_fill
                    elif status_enum == "INFO":
                        cell.fill = info_fill
                    elif status_enum == "PASS":
                        cell.fill = pass_fill

        # Phase 0: Heading Confidence column (26) - highlight when < 0.5
        self._apply_heading_confidence_format(ws)

    def _apply_heading_confidence_format(self, ws) -> None:
        """Phase 0: Conditional format for Heading Confidence (column 26) when value < 0.5."""
        from openpyxl.formatting.rule import CellIsRule
        from openpyxl.styles import PatternFill

        max_row = ws.max_row
        if max_row < 2:
            return
        col26_range = f"Z2:Z{max_row}"
        low_conf_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        rule = CellIsRule(
            operator="lessThan",
            formula=["0.5"],
            fill=low_conf_fill,
        )
        ws.conditional_formatting.add(col26_range, rule)

    def write_executive_summary(self, wb: Workbook, results: List[Dict]) -> None:
        """
        SCRUM-7: Write Executive Summary sheet with high-level metrics and top findings.

        Args:
            wb: Excel workbook
            results: List of validation results
        """
        ws = wb.create_sheet(title="Executive Summary", index=0)

        # Header
        ws["A1"] = "EXECUTIVE SUMMARY"
        ws["A1"].font = _font_with(ws["A1"].font, bold=True, size=16)

        # Metrics section
        ws["A3"] = "Metrics"
        ws["A3"].font = _font_with(ws["A3"].font, bold=True)

        # Count by status
        total = len(results)
        passed = sum(1 for r in results if r.get("status_enum") == "PASS")
        failed = sum(1 for r in results if r.get("status_enum") == "FAIL")
        warnings = sum(1 for r in results if r.get("status_enum") == "WARN")

        # Row 5: Metrics
        ws["A5"] = "Total Tables"
        ws["B5"] = total
        ws["C5"] = "Passed"
        ws["D5"] = passed
        ws["E5"] = "Failed"
        ws["F5"] = failed
        ws["G5"] = "Warnings"
        ws["H5"] = warnings

        # Overall Assessment
        ws["A7"] = "Overall Assessment"
        ws["A7"].font = _font_with(ws["A7"].font, bold=True)
        if failed > 0:
            overall = "RED"
        elif warnings > 0:
            overall = "YELLOW"
        else:
            overall = "GREEN"
        ws["B7"] = overall
        if overall == "RED":
            ws["B7"].fill = RED_FILL
        elif overall == "YELLOW":
            ws["B7"].fill = INFO_FILL
        else:
            ws["B7"].fill = GREEN_FILL

        # Top 10 Findings Table
        ws["A9"] = "Top 10 Findings (by Severity)"
        ws["A9"].font = _font_with(ws["A9"].font, bold=True)

        # Headers (Row 10)
        ws["A10"] = "Table Name"
        ws["B10"] = "Status"
        ws["C10"] = "Severity"
        ws["D10"] = "Max Diff"
        ws["E10"] = "Root Cause"
        for col in range(1, 6):
            ws.cell(row=10, column=col).font = _font_with(
                ws.cell(row=10, column=col).font, bold=True
            )

        # Filter and sort findings (FAIL/WARN and tool/data statuses, sorted by severity)
        findings = [
            r
            for r in results
            if r.get("status_enum")
            in ["FAIL", "WARN", "FAIL_TOOL_EXTRACT", "FAIL_TOOL_LOGIC", "FAIL_DATA"]
        ]
        # Sort by severity (HIGH > MEDIUM > LOW)
        severity_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
        findings.sort(
            key=lambda x: (
                severity_order.get(x.get("severity", "LOW"), 0),
                x.get("context", {}).get("max_diff", 0),
            ),
            reverse=True,
        )
        top_10 = findings[:10]

        # Write findings (starting row 11)
        for idx, finding in enumerate(top_10, start=11):
            heading = finding.get("context", {}).get("heading", "Unknown")
            ws.cell(row=idx, column=1, value=heading)
            ws.cell(row=idx, column=2, value=finding.get("status_enum", "UNKNOWN"))
            ws.cell(row=idx, column=3, value=finding.get("severity", "LOW"))
            context = finding.get("context", {}) or {}
            max_diff = context.get("max_diff")
            if max_diff is None:
                # Fallback: derive from evidence diffs if present
                evidence = finding.get("evidence")
                if isinstance(evidence, list):
                    diffs: List[float] = []
                    for item in evidence:
                        if not isinstance(item, dict):
                            continue
                        diff_val = item.get("diff")
                        if diff_val is None:
                            continue
                        try:
                            diffs.append(abs(float(diff_val)))
                        except Exception:
                            continue
                    if diffs:
                        max_diff = max(diffs)

            ws.cell(row=idx, column=4, value=max_diff if max_diff is not None else "")
            ws.cell(row=idx, column=5, value=finding.get("root_cause", "general"))

            # Apply status color
            status_enum = finding.get("status_enum", "")
            if status_enum == "FAIL":
                ws.cell(row=idx, column=2).fill = RED_FILL
            elif status_enum == "WARN":
                ws.cell(row=idx, column=2).fill = INFO_FILL

    def write_focus_list(
        self, wb: Workbook, results: List[Dict], telemetry=None
    ) -> None:
        """
        SCRUM-7: Write Focus List sheet with triage-ready findings sorted by severity.
        Phase 6: Extractor Engine, Quality Score, Failure Reason Code, run_id.

        Args:
            wb: Excel workbook
            results: List of validation results
            telemetry: Optional TelemetryCollector for run_id
        """
        ws = wb.create_sheet(title="Focus List", index=1)

        run_id = ""
        if telemetry is not None and getattr(telemetry, "run_telemetry", None):
            run_id = getattr(telemetry.run_telemetry, "run_id", "") or ""

        # Headers (Row 1) - SCRUM-7/8: Includes Review Status column; Phase 6: L–O; P4.2: P–T diagnostic
        headers = [
            "Table Name",  # A
            "Status",  # B
            "Severity",  # C
            "Confidence",  # D
            "Issue Description",  # E
            "Max Diff",  # F
            "Jump",  # G (hyperlink column)
            "Owner (user input)",  # H
            "Review Status",  # I (SCRUM-8: with data validation)
            "Comment (user input)",  # J
            "table_id",  # K (hidden)
            "Extractor Engine",  # L Phase 6
            "Quality Score",  # M
            "Failure Reason Code",  # N
            "run_id",  # O
            "Assertions Count",  # P P4.2
            "Numeric Evidence",  # Q
            "CY Column",  # R
            "PY Column",  # S
            "Reason Code",  # T
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = _font_with(cell.font, bold=True)

        # Hide column K (table_id)
        ws.column_dimensions["K"].hidden = True

        # Filter findings (FAIL/WARN and tool/data statuses)
        findings = [
            r
            for r in results
            if r.get("status_enum")
            in ["FAIL", "WARN", "FAIL_TOOL_EXTRACT", "FAIL_TOOL_LOGIC", "FAIL_DATA"]
        ]

        # Sort by severity (HIGH > MEDIUM > LOW)
        severity_order = {"HIGH": 3, "MEDIUM": 2, "LOW": 1}
        findings.sort(
            key=lambda x: (
                severity_order.get(x.get("severity", "LOW"), 0),
                x.get("context", {}).get("max_diff", 0),
            ),
            reverse=True,
        )

        # Write findings (starting row 2)
        for row_idx, finding in enumerate(findings, start=2):
            # SCRUM-8: Use heading from context.heading, not Unknown
            heading = (
                finding.get("context", {}).get("heading")
                or finding.get("context", {}).get("table_name")
                or "Unknown"
            )
            table_id = finding.get("table_id")
            status_enum = finding.get("status_enum", "UNKNOWN")
            # SCRUM-8: Ensure severity is not defaulting to MEDIUM - use calculated value or infer
            severity = finding.get("severity")
            if not severity or severity == "MEDIUM":
                # Recalculate if missing or default
                rule_id = finding.get("rule_id", "UNKNOWN")
                max_diff = finding.get("context", {}).get("max_diff", 0.0)
                is_skipped = (
                    "skip" in finding.get("status", "").lower()
                    or finding.get("status_enum") == "INFO"
                )
                # Use concrete validator to calculate severity
                from ..core.validators.generic_validator import GenericTableValidator

                temp_validator = GenericTableValidator()
                severity = temp_validator._calculate_severity(
                    rule_id, max_diff, is_skipped
                )

            confidence = finding.get("confidence", "MEDIUM")
            # SCRUM-8: Ensure root_cause is not defaulting to General
            root_cause = finding.get("root_cause")
            if not root_cause or root_cause == "general":
                # Recalculate if missing or default
                rule_id = finding.get("rule_id", "UNKNOWN")
                context = finding.get("context", {})
                from ..core.validators.generic_validator import GenericTableValidator

                temp_validator = GenericTableValidator()
                root_cause = temp_validator._infer_root_cause(rule_id, context)

            # Extract max_diff from multiple sources (SCRUM-8)
            max_diff = None
            context = finding.get("context", {})
            if "max_diff" in context:
                max_diff = context["max_diff"]
            elif finding.get("marks"):
                # Extract from marks
                diffs = [
                    abs(float(m.get("diff", 0)))
                    for m in finding.get("marks", [])
                    if m.get("diff") is not None
                ]
                if diffs:
                    max_diff = max(diffs)
            elif finding.get("evidence"):
                # Extract from evidence
                diffs = [
                    abs(float(e.get("diff", 0)))
                    for e in finding.get("evidence", [])
                    if e.get("diff") is not None
                ]
                if diffs:
                    max_diff = max(diffs)
            else:
                # Parse from status message
                status = finding.get("status", "")
                match = re.search(r"Sai lệch\s*[=:]\s*([\d,.-]+)", status)
                if match:
                    with contextlib.suppress(ValueError, TypeError):
                        max_diff = abs(float(match.group(1).replace(",", "")))

            # Write columns
            ws.cell(row=row_idx, column=1, value=heading)  # Table Name
            ws.cell(row=row_idx, column=2, value=status_enum)  # Status
            ws.cell(row=row_idx, column=3, value=severity)  # Severity
            ws.cell(row=row_idx, column=4, value=confidence)  # Confidence
            ws.cell(
                row=row_idx, column=5, value=finding.get("status", "")
            )  # Issue Description

            # SCRUM-8: Max Diff - format as number, leave blank if unavailable (not 0)
            max_diff_cell = ws.cell(
                row=row_idx,
                column=6,
                value=max_diff if max_diff is not None and max_diff != 0 else None,
            )
            if max_diff is not None and max_diff != 0:
                max_diff_cell.number_format = "#,##0"  # Format to prevent ######

            # SCRUM-8: Review Status (column I) defaults to "New"
            ws.cell(row=row_idx, column=9, value="New")  # Review Status

            # Comment (column J) is user input; initialize empty
            ws.cell(row=row_idx, column=10, value="")
            ws.cell(row=row_idx, column=11, value=table_id)  # table_id (hidden)
            # Phase 6: Extractor Engine, Quality Score, Failure Reason Code, run_id
            ctx = finding.get("context") or {}
            ws.cell(row=row_idx, column=12, value=ctx.get("extractor_engine", ""))
            qs = ctx.get("quality_score")
            ws.cell(row=row_idx, column=13, value=qs)
            ws.cell(row=row_idx, column=14, value=ctx.get("failure_reason_code", ""))
            ws.cell(row=row_idx, column=15, value=run_id)
            # P4.2: Diagnostic columns
            ws.cell(
                row=row_idx,
                column=16,
                value=finding.get("assertions_count", ctx.get("assertions_count")),
            )
            num_ev = ctx.get("numeric_evidence_score")
            ws.cell(
                row=row_idx,
                column=17,
                value=round(float(num_ev), 4) if num_ev is not None else None,
            )
            cy_col = ctx.get("cy_column")
            if cy_col is None and isinstance(ctx.get("amount_columns"), (list, tuple)):
                ac = ctx.get("amount_columns")
                cy_col = ac[0] if ac else None
            ws.cell(
                row=row_idx,
                column=18,
                value=sanitize_excel_value(cy_col if cy_col is not None else ""),
            )
            py_col = ctx.get("py_column")
            ac = ctx.get("amount_columns")
            if py_col is None and isinstance(ac, (list, tuple)) and len(ac) > 1:
                py_col = ac[1]
            ws.cell(
                row=row_idx,
                column=19,
                value=sanitize_excel_value(py_col if py_col is not None else ""),
            )
            reason_code = (
                ctx.get("reason_code")
                or finding.get("failure_reason_code")
                or ctx.get("failure_reason_code")
                or ""
            )
            ws.cell(row=row_idx, column=20, value=sanitize_excel_value(reason_code))

            # Apply status color to Status column (include FAIL_TOOL_EXTRACT, FAIL_TOOL_LOGIC, FAIL_DATA)
            if status_enum in (
                "FAIL",
                "FAIL_TOOL_EXTRACT",
                "FAIL_TOOL_LOGIC",
                "FAIL_DATA",
            ):
                ws.cell(row=row_idx, column=2).fill = RED_FILL
            elif status_enum == "WARN":
                ws.cell(row=row_idx, column=2).fill = INFO_FILL

            # Apply severity color to Severity column
            if severity == "HIGH":
                ws.cell(row=row_idx, column=3).fill = RED_FILL
            elif severity == "MEDIUM":
                ws.cell(row=row_idx, column=3).fill = INFO_FILL
            else:
                ws.cell(row=row_idx, column=3).fill = GREEN_FILL

            # SCRUM-8: Hyperlink in Jump column (G/7) - ensure ALL rows have hyperlinks
            if table_id and table_id in self.anchor_map:
                anchor_row = self.anchor_map[table_id]
                jump_cell = ws.cell(row=row_idx, column=7, value="Jump to Table")
                jump_cell.hyperlink = f"#'FS casting'!A{anchor_row}"
                jump_cell.style = "Hyperlink"
            else:
                # SCRUM-8: Diagnostic flag for missing anchor
                ws.cell(row=row_idx, column=7, value="(missing anchor)")
                finding["context"]["anchor_missing"] = True

            # Load previous triage data if available (SCRUM-8)
            if self._previous_triage_data:
                triage_key = (
                    f"{table_id}|{finding.get('rule_id', 'UNKNOWN')}"
                    if table_id
                    else f"{heading}|{finding.get('rule_id', 'UNKNOWN')}"
                )
                prev_data = self._previous_triage_data.get(triage_key, {})
                if prev_data:
                    ws.cell(
                        row=row_idx, column=8, value=prev_data.get("owner", "")
                    )  # Owner (H)
                    ws.cell(
                        row=row_idx,
                        column=9,
                        value=prev_data.get("review_status", "New"),
                    )  # Review Status (I)
                    ws.cell(
                        row=row_idx, column=10, value=prev_data.get("comment", "")
                    )  # Comment (J)

        # Set column widths
        ws.column_dimensions["A"].width = 30  # Table Name
        ws.column_dimensions["B"].width = 12  # Status
        ws.column_dimensions["C"].width = 12  # Severity
        ws.column_dimensions["D"].width = 12  # Confidence
        ws.column_dimensions["E"].width = 50  # Issue Description
        ws.column_dimensions["F"].width = 18  # Max Diff (increased to prevent ######)
        ws.column_dimensions["G"].width = 15  # Jump
        ws.column_dimensions["H"].width = 20  # Owner
        ws.column_dimensions["I"].width = 18  # Review Status
        ws.column_dimensions["J"].width = 40  # Comment
        ws.column_dimensions["L"].width = 18  # Extractor Engine
        ws.column_dimensions["M"].width = 14  # Quality Score
        ws.column_dimensions["N"].width = 22  # Failure Reason Code
        ws.column_dimensions["O"].width = 36  # run_id
        ws.column_dimensions["P"].width = 16  # Assertions Count
        ws.column_dimensions["Q"].width = 16  # Numeric Evidence
        ws.column_dimensions["R"].width = 14  # CY Column
        ws.column_dimensions["S"].width = 14  # PY Column
        ws.column_dimensions["T"].width = 20  # Reason Code

        # SCRUM-8: Add data validation for Review Status column (I)
        # Only add if there are data rows (max_row > 1, since row 1 is header)
        if ws.max_row > 1:
            from openpyxl.worksheet.datavalidation import DataValidation

            review_status_options = [
                "New",
                "In Review",
                "Confirmed Issue",
                "False Positive",
                "Resolved",
            ]
            dv = DataValidation(
                type="list",
                formula1=f'"{",".join(review_status_options)}"',
                allow_blank=False,
            )
            dv.add(
                f"I2:I{ws.max_row}"
            )  # Apply to all data rows in Review Status column
            ws.add_data_validation(dv)

    def write_telemetry_sheet(
        self,
        wb: Workbook,
        telemetry,
        *,
        skipped_footer_signature_count: int = 0,
    ) -> None:
        """
        E3: Write telemetry data to Run metadata sheet.

        Args:
            wb: Excel workbook
            telemetry: TelemetryCollector instance
            skipped_footer_signature_count: Count of footer/signature artifacts excluded from output (T1)
        """
        from ..utils.telemetry_collector import TelemetryCollector

        if not isinstance(telemetry, TelemetryCollector):
            return

        ws = wb.create_sheet(title="Run metadata")

        # Header
        ws["A1"] = "RUN METADATA"
        ws["A1"].font = _font_with(ws["A1"].font, bold=True, size=16)

        # Build info
        ws["A3"] = "Build Information"
        ws["A3"].font = _font_with(ws["A3"].font, bold=True)
        ws["A4"] = "Tool Version"
        ws["B4"] = telemetry.run_telemetry.tool_version
        ws["A5"] = "Git Commit"
        ws["B5"] = telemetry.run_telemetry.git_commit_hash
        ws["A6"] = "Run Timestamp"
        ws["B6"] = telemetry.run_telemetry.run_timestamp
        ws["A7"] = "Run ID"
        ws["B7"] = getattr(telemetry.run_telemetry, "run_id", "") or ""

        # Performance metrics
        ws["A8"] = "Performance Metrics"
        ws["A8"].font = _font_with(ws["A8"].font, bold=True)
        summary = telemetry.get_summary()
        ws["A9"] = "Total Runtime (ms)"
        ws["B9"] = summary.get("total_runtime_ms", 0)
        ws["A10"] = "Table Count"
        ws["B10"] = summary.get("table_count", 0)
        ws["A11"] = "Avg Table Runtime (ms)"
        ws["B11"] = summary.get("avg_table_runtime_ms", 0)
        ws["A12"] = "Total Rows Processed"
        ws["B12"] = summary.get("total_rows_processed", 0)
        ws["A13"] = "Total Cells Processed"
        ws["B13"] = summary.get("total_cells_processed", 0)
        ws["A14"] = "Skipped Footer/Signature Count"
        ws["B14"] = skipped_footer_signature_count

        # Tables by status
        ws["A15"] = "Tables by Status"
        ws["A15"].font = _font_with(ws["A15"].font, bold=True)
        row = 16
        for status, count in summary.get("tables_by_status", {}).items():
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Tables by validator
        ws["A20"] = "Tables by Validator"
        ws["A20"].font = _font_with(ws["A20"].font, bold=True)
        row = 21
        for validator, count in summary.get("tables_by_validator", {}).items():
            ws.cell(row=row, column=1, value=validator)
            ws.cell(row=row, column=2, value=count)
            row += 1

        # Phase 6: Per-Table Extraction traceability
        start_row = row + 2
        ws.cell(row=start_row, column=1, value="Per-Table Extraction")
        ws.cell(row=start_row, column=1).font = _font_with(
            ws.cell(row=start_row, column=1).font, bold=True
        )
        start_row += 1
        headers = [
            "Table Index",
            "Heading",
            "Heading Source",
            "Heading Confidence",
            "Extractor Engine",
            "Quality Score",
            "Failure Reason Code",
            "Totals Candidates",
            "Totals Equations Solved",
            "Classifier Type",
            "Classifier Confidence",
            "Classifier Reason",
            "Extractor Usable Reason",
            "Assertions Count",
        ]
        for c, h in enumerate(headers, start=1):
            ws.cell(row=start_row, column=c, value=h)
            ws.cell(row=start_row, column=c).font = _font_with(
                ws.cell(row=start_row, column=c).font, bold=True
            )
        start_row += 1
        for t in getattr(telemetry.run_telemetry, "tables", []) or []:
            ws.cell(row=start_row, column=1, value=t.table_index)
            ws.cell(row=start_row, column=2, value=t.heading or "")
            ws.cell(
                row=start_row, column=3, value=getattr(t, "heading_source", None) or ""
            )
            ws.cell(
                row=start_row, column=4, value=getattr(t, "heading_confidence", None)
            )
            ws.cell(row=start_row, column=5, value=t.extractor_engine or "")
            ws.cell(row=start_row, column=6, value=t.quality_score)
            ws.cell(row=start_row, column=7, value=t.failure_reason_code or "")
            ws.cell(row=start_row, column=8, value=t.totals_candidates_found)
            ws.cell(row=start_row, column=9, value=t.totals_equations_solved)
            ws.cell(
                row=start_row,
                column=10,
                value=getattr(t, "classifier_primary_type", None) or "",
            )
            ws.cell(
                row=start_row,
                column=11,
                value=getattr(t, "classifier_confidence", None),
            )
            ws.cell(
                row=start_row,
                column=12,
                value=getattr(t, "classifier_reason", None) or "",
            )
            ws.cell(
                row=start_row,
                column=13,
                value=getattr(t, "extractor_usable_reason", None) or "",
            )
            ws.cell(
                row=start_row,
                column=14,
                value=getattr(t, "assertions_count", None),
            )
            start_row += 1

        # R1: Baseline aggregates (invariant histogram, OOXML adoption, auditability)
        tables_list = getattr(telemetry.run_telemetry, "tables", []) or []
        total_tables = len(tables_list)
        tables_with_invariants_failed = sum(
            1
            for t in tables_list
            if getattr(t, "invariants_failed", None)
            and len(t.invariants_failed or []) > 0
        )
        ooxml_used = sum(
            1
            for t in tables_list
            if getattr(t, "engine_attempts", None)
            and any("ooxml" in (x or "").lower() for x in (t.engine_attempts or []))
        )
        tables_with_grid = sum(
            1 for t in tables_list if getattr(t, "grid_cols_expected", None) is not None
        )
        grid_match = sum(
            1
            for t in tables_list
            if getattr(t, "grid_cols_expected", None) is not None
            and getattr(t, "grid_cols_built", None) == t.grid_cols_expected
        )
        agg_row = start_row + 2
        ws.cell(row=agg_row, column=1, value="R1 Baseline Aggregates")
        ws.cell(row=agg_row, column=1).font = _font_with(
            ws.cell(row=agg_row, column=1).font, bold=True
        )
        agg_row += 1
        ws.cell(row=agg_row, column=1, value="Tables with invariant violations")
        ws.cell(row=agg_row, column=2, value=tables_with_invariants_failed)
        agg_row += 1
        ws.cell(row=agg_row, column=1, value="OOXML adoption (tables)")
        ws.cell(
            row=agg_row,
            column=2,
            value=f"{ooxml_used} / {total_tables}" if total_tables else "0 / 0",
        )
        ws.cell(
            row=agg_row,
            column=3,
            value=round(100.0 * ooxml_used / total_tables, 1) if total_tables else 0,
        )
        agg_row += 1
        ws.cell(row=agg_row, column=1, value="Auditability rate (grid match)")
        ws.cell(
            row=agg_row,
            column=2,
            value=f"{grid_match} / {tables_with_grid}" if tables_with_grid else "0 / 0",
        )
        ws.cell(
            row=agg_row,
            column=3,
            value=(
                round(100.0 * grid_match / tables_with_grid, 1)
                if tables_with_grid
                else 0
            ),
        )
        # Phase 0: Aggregates by heading_source, classifier_type, extractor_engine
        agg_row += 2
        ws.cell(row=agg_row, column=1, value="Phase 0 Aggregates")
        ws.cell(row=agg_row, column=1).font = _font_with(
            ws.cell(row=agg_row, column=1).font, bold=True
        )
        agg_row += 1
        from collections import Counter

        by_heading = Counter(
            getattr(t, "heading_source", None) or "(none)" for t in tables_list
        )
        ws.cell(row=agg_row, column=1, value="By Heading Source")
        agg_row += 1
        for k, v in sorted(by_heading.items()):
            ws.cell(row=agg_row, column=1, value=k)
            ws.cell(row=agg_row, column=2, value=v)
            agg_row += 1
        agg_row += 1
        by_classifier = Counter(
            getattr(t, "classifier_primary_type", None) or "(none)" for t in tables_list
        )
        ws.cell(row=agg_row, column=1, value="By Classifier Type")
        agg_row += 1
        for k, v in sorted(by_classifier.items()):
            ws.cell(row=agg_row, column=1, value=k)
            ws.cell(row=agg_row, column=2, value=v)
            agg_row += 1
        agg_row += 1
        by_engine = Counter(
            getattr(t, "extractor_engine", None) or "(none)" for t in tables_list
        )
        ws.cell(row=agg_row, column=1, value="By Extractor Engine")
        agg_row += 1
        for k, v in sorted(by_engine.items()):
            ws.cell(row=agg_row, column=1, value=k)
            ws.cell(row=agg_row, column=2, value=v)
            agg_row += 1
