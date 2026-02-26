#!/usr/bin/env python3
"""
Phase 1: Shortlist FAIL/INFO from audit XLSX for false-fail reduction.

Reads summary sheet from one or more XLSX (e.g. from run_regression_2docs).
Filters FAIL and INFO; groups by validator_type, rule_id, failure_reason_code, table_id.
Outputs ~10 case shortlist with root-cause hypothesis to reports/shortlist_fail_info.md (and .json).
"""

import argparse
import json
import sys
from pathlib import Path

_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

try:
    import openpyxl
except ImportError:
    openpyxl = None

SHEET_NAME = "Tổng hợp kiểm tra"


def _read_summary_rows(xlsx_path: Path) -> list[dict]:
    """Read summary sheet rows as list of dicts (header from first row)."""
    if not openpyxl:
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet_alt = None
    if SHEET_NAME not in wb.sheetnames:
        for name in wb.sheetnames:
            if "summary" in name.lower() or "tong hop" in name.lower():
                sheet_alt = name
                break
        if sheet_alt is None and wb.sheetnames:
            sheet_alt = wb.sheetnames[0]
        ws = wb[sheet_alt] if sheet_alt else None
    else:
        ws = wb[SHEET_NAME]
    if ws is None:
        wb.close()
        return []
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [
        str(c).strip() if c is not None else f"Col{i}" for i, c in enumerate(rows[0])
    ]
    out = []
    for row in rows[1:]:
        d = {}
        for i, v in enumerate(row):
            if i < len(header) and header[i]:
                d[header[i]] = v
        out.append(d)
    return out


def _status_enum(row: dict) -> str:
    v = row.get("Status Enum") or row.get("status_enum")
    if v is None:
        for k in row:
            if "status" in k.lower() and "enum" in k.lower():
                v = row.get(k)
                break
    return str(v).strip().upper() if v is not None else ""


def _is_fail_or_info(row: dict) -> bool:
    s = _status_enum(row)
    return s == "FAIL" or s == "INFO" or s.startswith("FAIL_") or s == "INFO_SKIPPED"


def _table_id(row: dict) -> str:
    return str(row.get("Table ID") or row.get("table_id") or "").strip() or row.get(
        "Tên bảng", ""
    )


def _hypothesis(validator_type: str, rule_id: str, failure_code: str) -> str:
    """Root-cause hypothesis by validator/rule pattern."""
    vt = (validator_type or "").lower()
    rid = (rule_id or "").upper()
    frc = (failure_code or "").upper()
    if "generic" in vt and ("COLUMN_TOTAL" in rid or "TOTAL" in frc):
        return "Total row not detected or wrong row; enable_generic_total_gate / tighten_total_row_keywords may help."
    if "generic" in vt and ("NO_EVIDENCE" in frc or "0 assertion" in str(frc)):
        return "No assertions run; treat_no_assertion_as_pass or eligibility gate."
    if "equity" in vt:
        return "Balance-at sum = 0 vs table value; equity_no_evidence_not_fail or header/column mapping."
    if "balance" in vt:
        return "Total row or amount column mismatch; safe_total_row_selection or routing gate."
    if "cash" in vt:
        return "Cross-table or period mismatch; cashflow_cross_table_context."
    return "Review validator logic and extraction quality for this table."


def shortlist_from_xlsx_paths(
    xlsx_paths: list[Path],
    max_cases: int = 10,
    output_dir: Path | None = None,
) -> dict:
    """
    Read XLSX, filter FAIL/INFO, group, produce shortlist with root-cause hypothesis.

    Returns dict: shortlist (list of case dicts), groups (by validator, rule_id, failure_code), stats.
    """
    all_rows = []
    for path in xlsx_paths:
        if not path.exists():
            continue
        for row in _read_summary_rows(path):
            row["_source_file"] = path.name
            if _is_fail_or_info(row):
                all_rows.append(row)

    # Group by (validator_type, rule_id, failure_reason_code) and collect table_ids
    groups = {}
    for row in all_rows:
        vt = row.get("Validator Type") or row.get("validator_type") or "?"
        rid = row.get("Rule ID") or row.get("rule_id") or "?"
        frc = row.get("Failure Reason Code") or row.get("failure_reason_code") or ""
        key = (vt, rid, frc)
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    # Build shortlist: one representative per group + table_id diversity, up to max_cases
    shortlist = []
    seen_table_ids = set()
    for (vt, rid, frc), rows in sorted(groups.items(), key=lambda x: -len(x[1])):
        if len(shortlist) >= max_cases:
            break
        # Pick first row that adds a new table_id or first row of group
        rep = None
        for r in rows:
            tid = _table_id(r)
            if tid and tid not in seen_table_ids:
                rep = r
                seen_table_ids.add(tid)
                break
        if rep is None:
            rep = rows[0]
        shortlist.append(
            {
                "table_id": _table_id(rep),
                "status": _status_enum(rep),
                "validator_type": vt,
                "rule_id": rid,
                "failure_reason_code": frc,
                "source_file": rep.get("_source_file", ""),
                "count_in_group": len(rows),
                "root_cause_hypothesis": _hypothesis(vt, rid, frc),
            }
        )

    stats = {
        "total_fail_info_rows": len(all_rows),
        "group_count": len(groups),
        "shortlist_count": len(shortlist),
    }
    return {
        "shortlist": shortlist,
        "groups": [
            {
                "validator_type": k[0],
                "rule_id": k[1],
                "failure_reason_code": k[2],
                "count": len(v),
                "sample_table_id": _table_id(v[0]) if v else "",
            }
            for k, v in sorted(groups.items(), key=lambda x: -len(x[1]))
        ],
        "stats": stats,
    }


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Phase 1: Shortlist FAIL/INFO from audit XLSX with root-cause hypothesis."
    )
    parser.add_argument(
        "xlsx_paths",
        nargs="*",
        type=Path,
        help="XLSX paths (if empty, glob reports/*.xlsx)",
    )
    parser.add_argument(
        "--max",
        type=int,
        default=10,
        help="Max shortlist cases (default 10)",
    )
    parser.add_argument(
        "-o",
        "--output-dir",
        type=Path,
        default=_project_root / "reports",
        help="Output directory for shortlist_fail_info.md and .json",
    )
    args = parser.parse_args()

    paths = list(args.xlsx_paths)
    if not paths:
        for p in (_project_root / "reports").glob("*.xlsx"):
            paths.append(p)
    if not paths:
        print(
            "No XLSX given and none in reports/. Run: python scripts/run_regression_2docs.py [doc1] [doc2]",
            file=sys.stderr,
        )
        return 1

    if not openpyxl:
        print("openpyxl required", file=sys.stderr)
        return 1

    out = shortlist_from_xlsx_paths(
        paths, max_cases=args.max, output_dir=args.output_dir
    )
    args.output_dir.mkdir(parents=True, exist_ok=True)

    json_path = args.output_dir / "shortlist_fail_info.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2, default=str)

    md_lines = [
        "# Shortlist FAIL/INFO (Phase 1)",
        "",
        f"Total FAIL/INFO rows: {out['stats']['total_fail_info_rows']}",
        f"Groups: {out['stats']['group_count']}",
        "",
        "## Shortlist (~10 cases with root-cause hypothesis)",
        "",
        "| # | table_id | status | validator_type | rule_id | failure_code | count | root_cause_hypothesis |",
        "|---|----------|--------|-----------------|--------|--------------|-------|------------------------|",
    ]
    for i, c in enumerate(out["shortlist"], 1):
        md_lines.append(
            f"| {i} | {c['table_id']} | {c['status']} | {c['validator_type']} | {c['rule_id']} | {c['failure_reason_code'][:20]} | {c['count_in_group']} | {c['root_cause_hypothesis'][:60]} |"
        )
    md_lines.extend(["", "## Groups (by validator, rule_id, failure_code)", ""])
    for g in out["groups"][:20]:
        md_lines.append(
            f"- {g['validator_type']} | {g['rule_id']} | {g['failure_reason_code'][:30]} | count={g['count']} | sample={g['sample_table_id']}"
        )
    md_path = args.output_dir / "shortlist_fail_info.md"
    md_path.write_text("\n".join(md_lines), encoding="utf-8")

    print(f"Wrote {json_path} and {md_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
