"""Enhanced parse script to extract ALL forensic fields from XLSX for audit."""

import json
import sys
from collections import defaultdict
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
import openpyxl  # noqa: E402


def parse_xlsx_forensic(filepath: Path) -> dict:
    """Parse XLSX with all forensic fields for forensic audit."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result = {
        "file": filepath.name,
        "sheets": wb.sheetnames,
        "summary_data": [],
        "all_headers": [],
    }

    # Find summary sheet
    summary_sheet = None
    for name in wb.sheetnames:
        if "Tổng hợp kiểm tra" in name or "Tong hop kiem tra" in name:
            summary_sheet = wb[name]
            result["summary_sheet_name"] = name
            break

    if not summary_sheet:
        for name in wb.sheetnames:
            if "summary" in name.lower() and "executive" not in name.lower():
                summary_sheet = wb[name]
                result["summary_sheet_name"] = name
                break

    if summary_sheet:
        # Read ALL headers from first row
        headers = []
        for cell in summary_sheet[1]:
            headers.append(
                str(cell.value).strip() if cell.value else f"Col_{cell.column}"
            )
        result["all_headers"] = headers

        # Read all data rows with ALL fields
        for row_num, row in enumerate(
            summary_sheet.iter_rows(min_row=2, max_row=200), start=2
        ):
            row_data = {"_row_num": row_num}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = cell.value
            if any(v for k, v in row_data.items() if k != "_row_num"):
                result["summary_data"].append(row_data)

    wb.close()
    return result


def analyze_forensic(all_data: list) -> dict:
    """Analyze all data with forensic detail."""
    status_counts = defaultdict(int)
    fail_tool_extract = []
    fail_warn = []
    pivot_total_row_method = defaultdict(lambda: {"FAIL": 0, "WARN": 0})
    pivot_engine_attempts = defaultdict(int)

    observability_gaps = {
        "table_id_empty": 0,
        "failure_reason_empty": 0,
        "extractor_engine_empty": 0,
        "total_row_method_empty": 0,
    }

    for row in all_data:
        # Normalize status
        status = None
        for key in ["Status Enum", "status_enum", "Status", "Trạng thái"]:
            if key in row and row[key]:
                status = str(row[key]).strip().upper()
                break

        if not status:
            continue

        status_counts[status] += 1

        # Extract normalized fields
        table_name = row.get("Tên bảng") or row.get("Col_1") or "?"
        validator = row.get("Validator Type") or row.get("validator_type") or "?"
        failure_reason = (
            row.get("Failure Reason Code") or row.get("failure_reason") or ""
        )
        quality_score = row.get("Quality Score") or row.get("quality_score") or ""
        extractor_engine = (
            row.get("Extractor Engine") or row.get("extractor_engine") or ""
        )
        total_row_method = (
            row.get("Total Row Method") or row.get("total_row_method") or ""
        )
        engine_attempts = row.get("Engine Attempts") or ""

        # Track observability gaps
        if not table_name or table_name == "?":
            observability_gaps["table_id_empty"] += 1
        if not failure_reason and status not in ("PASS", "INFO"):
            observability_gaps["failure_reason_empty"] += 1
        if not extractor_engine:
            observability_gaps["extractor_engine_empty"] += 1
        if not total_row_method and status in ("FAIL", "WARN"):
            observability_gaps["total_row_method_empty"] += 1

        # Build inventories
        if status == "FAIL_TOOL_EXTRACT":
            fail_tool_extract.append(
                {
                    "row": row.get("_row_num"),
                    "table_name": table_name[:50] if table_name else "?",
                    "validator": validator,
                    "extractor_engine": extractor_engine or "?",
                    "engine_attempts": engine_attempts or "?",
                    "quality_score": quality_score,
                    "failure_reason": failure_reason or "?",
                }
            )
            # Pivot by engine_attempts
            ea_key = str(engine_attempts) if engine_attempts else "UNKNOWN"
            pivot_engine_attempts[ea_key] += 1

        elif status in ("FAIL", "WARN"):
            fail_warn.append(
                {
                    "row": row.get("_row_num"),
                    "table_name": table_name[:50] if table_name else "?",
                    "status": status,
                    "validator": validator,
                    "failure_reason": failure_reason or "?",
                    "total_row_method": total_row_method or "?",
                    "quality_score": quality_score,
                }
            )
            # Pivot by total_row_method
            trm_key = str(total_row_method) if total_row_method else "UNKNOWN"
            pivot_total_row_method[trm_key][status] += 1

    return {
        "status_counts": dict(status_counts),
        "total": sum(status_counts.values()),
        "fail_tool_extract_inventory": fail_tool_extract,
        "fail_warn_inventory": fail_warn,
        "pivot_total_row_method": dict(pivot_total_row_method),
        "pivot_engine_attempts": dict(pivot_engine_attempts),
        "observability_gaps": observability_gaps,
    }


def main():
    results_dir = Path("results")
    xlsx_files = list(results_dir.glob("*_output.xlsx"))

    combined_data = []
    file_results = []

    for xlsx_file in xlsx_files:
        print(f"\n=== {xlsx_file.name} ===")
        data = parse_xlsx_forensic(xlsx_file)
        file_results.append(data)

        print(f"Sheets: {data['sheets']}")
        print(f"Summary sheet: {data.get('summary_sheet_name', 'NOT FOUND')}")
        print(f"Headers ({len(data['all_headers'])}): {data['all_headers']}")
        print(f"Rows: {len(data['summary_data'])}")

        for row in data["summary_data"]:
            row["_source_file"] = xlsx_file.name
            combined_data.append(row)

    print("\n" + "=" * 60)
    print("FORENSIC ANALYSIS")
    print("=" * 60)

    analysis = analyze_forensic(combined_data)

    print(f"\nStatus Counts: {analysis['status_counts']}")
    print(f"Total: {analysis['total']}")

    print("\n--- Observability Gaps ---")
    for gap, count in analysis["observability_gaps"].items():
        print(f"  {gap}: {count}")

    print("\n--- Pivot: FAIL/WARN by Total Row Method ---")
    for method, counts in analysis["pivot_total_row_method"].items():
        print(f"  {method}: FAIL={counts['FAIL']}, WARN={counts['WARN']}")

    print("\n--- Pivot: FAIL_TOOL_EXTRACT by Engine Attempts ---")
    for ea, count in analysis["pivot_engine_attempts"].items():
        print(f"  {ea}: {count}")

    print(
        f"\n--- FAIL_TOOL_EXTRACT Inventory ({len(analysis['fail_tool_extract_inventory'])}) ---"
    )
    for item in analysis["fail_tool_extract_inventory"][:10]:
        print(
            f"  Row {item['row']}: engine={item['extractor_engine']}, attempts={item['engine_attempts']}, score={item['quality_score']}"
        )

    print(f"\n--- FAIL/WARN Inventory ({len(analysis['fail_warn_inventory'])}) ---")
    for item in analysis["fail_warn_inventory"][:10]:
        print(
            f"  Row {item['row']}: {item['status']} | {item['validator']} | method={item['total_row_method']}"
        )

    # Save results
    output = {
        "file_results": [
            {
                "file": f["file"],
                "headers": f["all_headers"],
                "row_count": len(f["summary_data"]),
            }
            for f in file_results
        ],
        "analysis": analysis,
    }
    with open("results/forensic_analysis.json", "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2, default=str)
    print("\nSaved: results/forensic_analysis.json")


if __name__ == "__main__":
    main()
