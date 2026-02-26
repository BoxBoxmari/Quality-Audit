#!/usr/bin/env python3
"""
P0: Aggregate FAIL/WARN from Quality Audit Excel outputs.

Reads sheet "Tổng hợp kiểm tra" from one or more XLSX files.
Groups by: validator_type, failure_reason_code, rule_id, extractor_engine,
total_row_method (if present). Outputs CSV and JSON to a given path prefix.
"""

import argparse
import csv
import json
import sys
from pathlib import Path

# Project root for imports when run as script
_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

try:
    import openpyxl
except ImportError:
    openpyxl = None


SHEET_NAME = "Tổng hợp kiểm tra"
STATUS_ENUM_COL = 3  # 1-based column C
FAIL_STATUSES = {"FAIL", "FAIL_TOOL_EXTRACT", "FAIL_VALIDATION"}
WARN_STATUSES = {"WARN"}

# Header indices (1-based) for "Tổng hợp kiểm tra" per excel_writer
HEADER_NAMES = [
    "Tên bảng",
    "Trạng thái kiểm tra",
    "Status Enum",
    "Status Category",
    "Rule ID",
    "Validator Type",
    "Extractor Engine",
    "Quality Score",
    "Failure Reason Code",
    "run_id",
    "Engine Attempts",
    "Invariants Failed",
    "Grid Cols Expected",
    "Grid Cols Built",
    "GridSpan Count",
    "vMerge Count",
]


def _read_summary_rows(xlsx_path: str) -> list[dict]:
    """Read summary sheet rows as list of dicts (header from first row)."""
    if not openpyxl:
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        return []
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []
    header = [
        str(c).strip() if c is not None else f"Col{i}" for i, c in enumerate(rows[0])
    ]
    out = []
    for row in rows[1:]:
        d = {}
        for i, v in enumerate(row):
            if i < len(header):
                d[header[i]] = v
        out.append(d)
    return out


def _status_enum(row: dict) -> str:
    """Get status enum from row (Status Enum or column C)."""
    v = row.get("Status Enum") or row.get("status_enum")
    if v is None:
        keys = [k for k in row if "status" in k.lower() and "enum" in k.lower()]
        if keys:
            v = row.get(keys[0])
    return str(v).strip().upper() if v is not None else ""


def _is_fail_or_warn(row: dict) -> bool:
    """
    Decide if a summary row should be counted as FAIL/WARN for aggregation.

    Default behaviour: include any row whose Status Enum is in FAIL_STATUSES/WARN_STATUSES.

    Special case (SCRUM-7 noise reduction):
    - Some tables are classified by GenericTableValidator with failure_reason_code
      'GenericTableValidator_VALIDATION' but the underlying check status is purely
      informational: "INFO: Bảng không bao gồm số/số tổng".
    - These are non-financial / narrative tables without numbers or totals and should
      not contribute to the global FAIL/WARN metrics.
    - We keep the original Excel output untouched, but exclude these rows from the
      aggregated failure report.
    """
    status_enum = _status_enum(row)

    # Noise gate: skip non-numeric/no-total tables that were reported as generic validation
    failure_reason = (
        row.get("Failure Reason Code") or row.get("failure_reason_code") or ""
    )
    status_text = str(row.get("Trạng thái kiểm tra") or row.get("status") or "")
    if (
        failure_reason == "GenericTableValidator_VALIDATION"
        and "Bảng không bao gồm số/số tổng" in status_text
    ):
        return False

    return status_enum in FAIL_STATUSES or status_enum in WARN_STATUSES


def _group_key(row: dict) -> tuple:
    """Key for grouping: validator_type, failure_reason_code, rule_id, extractor_engine, total_row_method."""
    return (
        (row.get("Validator Type") or row.get("validator_type") or ""),
        (row.get("Failure Reason Code") or row.get("failure_reason_code") or ""),
        (row.get("Rule ID") or row.get("rule_id") or ""),
        (row.get("Extractor Engine") or row.get("extractor_engine") or ""),
        (row.get("Total Row Method") or row.get("total_row_method") or ""),
    )


def aggregate_from_xlsx_paths(xlsx_paths: list[str], output_path_prefix: str) -> None:
    """
    Read all XLSX, collect FAIL/WARN rows, group by key, write CSV and JSON.

    output_path_prefix: e.g. reports/aggregate_failures -> writes
    reports/aggregate_failures.csv and reports/aggregate_failures.json
    """
    all_rows = []
    for path in xlsx_paths:
        for row in _read_summary_rows(path):
            row["_source_file"] = Path(path).name
            if _is_fail_or_warn(row):
                all_rows.append(row)

    # Group by (validator_type, failure_reason_code, rule_id, extractor_engine, total_row_method)
    groups = {}
    for row in all_rows:
        k = _group_key(row)
        if k not in groups:
            groups[k] = []
        groups[k].append(row)

    # Build aggregate records for CSV/JSON
    agg_records = []
    for (vt, frc, rid, ee, trm), rows in sorted(groups.items()):
        agg_records.append(
            {
                "validator_type": vt,
                "failure_reason_code": frc,
                "rule_id": rid,
                "extractor_engine": ee,
                "total_row_method": trm,
                "count": len(rows),
                "sources": list({r.get("_source_file", "") for r in rows}),
            }
        )

    prefix = Path(output_path_prefix)
    prefix.parent.mkdir(parents=True, exist_ok=True)

    csv_path = prefix.with_suffix(".csv")
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=[
                "validator_type",
                "failure_reason_code",
                "rule_id",
                "extractor_engine",
                "total_row_method",
                "count",
                "sources",
            ],
        )
        w.writeheader()
        for r in agg_records:
            row = dict(r)
            row["sources"] = ";".join(row["sources"])
            w.writerow(row)

    json_path = prefix.with_suffix(".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(
            {
                "total_fail_warn_rows": len(all_rows),
                "group_count": len(agg_records),
                "groups": agg_records,
            },
            f,
            indent=2,
            ensure_ascii=False,
        )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Aggregate FAIL/WARN from Quality Audit Excel outputs."
    )
    parser.add_argument(
        "xlsx_paths",
        nargs="+",
        type=Path,
        help="One or more XLSX paths (with sheet 'Tổng hợp kiểm tra')",
    )
    parser.add_argument(
        "-o",
        "--output-prefix",
        type=Path,
        default=_project_root / "reports" / "aggregate_failures",
        help="Output path prefix for .csv and .json (default: reports/aggregate_failures)",
    )
    args = parser.parse_args()

    if not openpyxl:
        print("openpyxl is required to read XLSX", file=sys.stderr)
        return 1

    # Expand globs in Python (Windows does not expand * on command line)
    resolved: list[Path] = []
    for p in args.xlsx_paths:
        if "*" in str(p):
            parent, name = p.parent, p.name
            for f in sorted(parent.glob(name)):
                if f.is_file() and f.suffix.lower() == ".xlsx":
                    resolved.append(f)
        elif p.is_file():
            resolved.append(p)
    if not resolved:
        print("No XLSX files found (paths or globs).", file=sys.stderr)
        return 1

    aggregate_from_xlsx_paths(
        [str(p) for p in resolved],
        str(args.output_prefix),
    )
    print(
        f"Wrote {args.output_prefix.with_suffix('.csv')} and {args.output_prefix.with_suffix('.json')}"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
