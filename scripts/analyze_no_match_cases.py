#!/usr/bin/env python3
"""Analyze safe_total_row_selection_no_match cases from XLSX reports."""

import sys
from pathlib import Path

_project_root = Path(__file__).resolve().parent.parent
if str(_project_root) not in sys.path:
    sys.path.insert(0, str(_project_root))

import openpyxl  # noqa: E402 (script: import after sys.path)


def analyze_no_match_cases(xlsx_path: Path):
    """Extract and analyze cases with safe_total_row_selection_no_match."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    # Find summary sheet
    summary_sheet = None
    for name in wb.sheetnames:
        if (
            "tong hop" in name.lower()
            or "summary" in name.lower()
            or "tổng hợp" in name.lower()
        ):
            summary_sheet = wb[name]
            break
    if not summary_sheet and wb.sheetnames:
        summary_sheet = wb[wb.sheetnames[0]]

    if not summary_sheet:
        print(f"No summary sheet found in {xlsx_path.name}")
        wb.close()
        return

    rows = list(summary_sheet.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return

    # Find header row
    header_row = None
    header_idx = None
    for i, row in enumerate(rows):
        row_str = [str(c).lower() if c else "" for c in row]
        if "total_row_method" in " ".join(row_str) or "status" in " ".join(row_str):
            header_row = row
            header_idx = i
            break

    if not header_row:
        print(f"No header row found in {xlsx_path.name}")
        wb.close()
        return

    # Map column names
    col_map = {}
    for i, col_name in enumerate(header_row):
        if col_name:
            col_map[str(col_name).strip().lower()] = i

    # Find cases with safe_total_row_selection_no_match
    no_match_cases = []
    for i, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        total_row_method = None
        status = None
        table_id = None
        validator_type = None
        rule_id = None
        failure_reason = None

        for key, idx in col_map.items():
            if idx < len(row):
                val = row[idx]
                if "total_row_method" in key:
                    total_row_method = val
                elif "status" in key and ("enum" in key or not status):
                    status = val
                elif "table" in key and "id" in key:
                    table_id = val
                elif "validator" in key and "type" in key:
                    validator_type = val
                elif "rule" in key and "id" in key:
                    rule_id = val
                elif "failure" in key and "reason" in key:
                    failure_reason = val

        if total_row_method == "safe_total_row_selection_no_match":
            no_match_cases.append(
                {
                    "row_num": i,
                    "table_id": table_id,
                    "status": status,
                    "validator_type": validator_type,
                    "rule_id": rule_id,
                    "failure_reason": failure_reason,
                    "source_file": xlsx_path.name,
                }
            )

    wb.close()

    print(f"\n=== {xlsx_path.name} ===")
    print(f"Found {len(no_match_cases)} safe_total_row_selection_no_match cases:")
    for case in no_match_cases:
        print(
            f"  Row {case['row_num']}: table_id={case['table_id']}, status={case['status']}, "
            f"validator={case['validator_type']}, rule={case['rule_id']}, "
            f"failure={case['failure_reason']}"
        )

    return no_match_cases


if __name__ == "__main__":

    xlsx_files = list(
        Path(_project_root / "reports").glob("fallback_no_amount_cols*.xlsx")
    )
    if not xlsx_files:
        print("No fallback_no_amount_cols XLSX files found")
        sys.exit(1)

    all_cases = []
    for xlsx_path in sorted(xlsx_files):
        cases = analyze_no_match_cases(xlsx_path)
        if cases:
            all_cases.extend(cases)

    print(f"\n=== Total: {len(all_cases)} safe_total_row_selection_no_match cases ===")
