"""Verify P0 columns are populated in XLSX output.

Usage (from repo root):
  python scripts/verify_p0_columns.py path/to/audit_output.xlsx
"""

import argparse
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl  # noqa: E402


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Verify P0 columns are populated in audit XLSX output."
    )
    parser.add_argument(
        "xlsx_path",
        type=str,
        help="Path to audit output XLSX (e.g. results/some_output.xlsx)",
    )
    return parser.parse_args()


args = parse_args()
wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

# Check 'Tong hop kiem tra' sheet for P0 columns
summary_sheet = wb["Tổng hợp kiểm tra"]

# Get headers from row 1
headers = [cell.value for cell in summary_sheet[1]]
print("=== XLSX Headers ===")
for i, h in enumerate(headers):
    safe_h = repr(h) if h else "None"
    print(f"{i + 1}: {safe_h}")

# Find new P0 columns
target_cols = [
    "Render First Rejection",
    "Mean Cell Confidence",
    "Token Coverage",
    "Excluded Columns",
]
print("\n=== P0 Columns in XLSX ===")
for tc in target_cols:
    found = [i + 1 for i, h in enumerate(headers) if h and tc.lower() in str(h).lower()]
    status = f"column(s) {found}" if found else "NOT FOUND"
    print(f"{tc}: {status}")

# Sample data from FAIL_TOOL_EXTRACT rows
print("\n=== FAIL_TOOL_EXTRACT Rows (Sample) ===")
status_col_idx = next((i for i, h in enumerate(headers) if h == "Status Enum"), None)
rf_rejection_idx = next(
    (i for i, h in enumerate(headers) if h and "Render First Rejection" in str(h)), None
)
confidence_idx = next(
    (i for i, h in enumerate(headers) if h and "Mean Cell Confidence" in str(h)), None
)
coverage_idx = next(
    (i for i, h in enumerate(headers) if h and "Token Coverage" in str(h)), None
)
excluded_idx = next(
    (i for i, h in enumerate(headers) if h and "Excluded Columns" in str(h)), None
)

print(
    f"Column indices: status={status_col_idx}, rf={rf_rejection_idx}, conf={confidence_idx}, cov={coverage_idx}, exc={excluded_idx}"
)

if status_col_idx is not None:
    fail_count = 0
    for row in summary_sheet.iter_rows(min_row=2, max_row=100):
        status = row[status_col_idx].value
        if status == "FAIL_TOOL_EXTRACT":
            fail_count += 1
            rf_val = (
                row[rf_rejection_idx].value if rf_rejection_idx is not None else "N/A"
            )
            conf_val = (
                row[confidence_idx].value if confidence_idx is not None else "N/A"
            )
            cov_val = row[coverage_idx].value if coverage_idx is not None else "N/A"
            exc_val = row[excluded_idx].value if excluded_idx is not None else "N/A"
            print(
                f"  Row {fail_count}: RF={rf_val}, Conf={conf_val}, Cov={cov_val}, Excluded={exc_val}"
            )
            if fail_count >= 3:
                break
    print(f"  (Total shown: {fail_count})")

# Also check total_row_method is populated
print("\n=== total_row_method Population ===")
total_method_idx = next(
    (i for i, h in enumerate(headers) if h and "Total Row Method" in str(h)), None
)
if total_method_idx is not None:
    methods = {}
    for row in summary_sheet.iter_rows(min_row=2, max_row=100):
        m = row[total_method_idx].value or "EMPTY"
        methods[m] = methods.get(m, 0) + 1
    print(f"  Distribution: {methods}")
else:
    print("  Total Row Method column NOT FOUND")

wb.close()
print("\n=== VERIFICATION COMPLETE ===")
